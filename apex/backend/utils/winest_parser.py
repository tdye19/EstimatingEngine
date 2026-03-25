"""WinEst file parser — handles .est (OLE2 native) and .xlsx WinEst exports.

Supports three intake scenarios:
  A. Native .est files (Microsoft Compound Document / OLE2 format via olefile)
  B. WinEst Excel Format 1 — Item/Description/Quantity/Unit/Labor Hours/Labor Rate/Material Cost/Total
  C. WinEst Excel Format 2 — WBS/CSI Code/Description/Qty/UOM/Crew Size/Productivity Rate/Hours/Cost

Returns a standardized result dict used by Agent 1 to populate APEX's internal format.
"""

import logging
from typing import Optional

logger = logging.getLogger("apex.winest_parser")

# Required column sets for each WinEst Excel export format
WINEST_FORMAT1_HEADERS = {
    "Item", "Description", "Quantity", "Unit",
    "Labor Hours", "Labor Rate", "Material Cost", "Total",
}
WINEST_FORMAT2_HEADERS = {
    "WBS", "CSI Code", "Description", "Qty",
    "UOM", "Crew Size", "Productivity Rate", "Hours", "Cost",
}


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_winest_file(file_path: str) -> dict:
    """Parse a WinEst .est or .xlsx file into structured line items.

    Auto-detects the format based on file extension and, for .xlsx files,
    column headers in the first row.

    Returns:
        {
            "success":         bool,
            "format_detected": str,   # "est_native" | "xlsx_format1" | "xlsx_format2"
                                      #   | "unknown_xlsx" | "unsupported"
            "line_items":      list[dict],
            "warnings":        list[str],
            "error":           str | None,
        }
    """
    ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""

    if ext == "est":
        return _parse_est_file(file_path)
    elif ext in ("xlsx", "xls"):
        return _parse_xlsx_file(file_path)
    else:
        return {
            "success": False,
            "format_detected": "unsupported",
            "line_items": [],
            "warnings": [],
            "error": f"Unsupported file extension: .{ext}",
        }


def is_winest_xlsx(file_path: str) -> bool:
    """Return True if an .xlsx file matches a known WinEst export format."""
    result = _parse_xlsx_file(file_path)
    return result.get("format_detected") in ("xlsx_format1", "xlsx_format2")


# ---------------------------------------------------------------------------
# Scenario A — native .est file (OLE2 / Compound Document)
# ---------------------------------------------------------------------------

_OLE_FALLBACK_ERROR = (
    "Unable to read this .est file. Please export your estimate from WinEst as "
    ".xlsx (File \u2192 Export \u2192 Excel) and upload the .xlsx file instead."
)

# Stream name fragments that typically contain estimate line-item data
_DATA_STREAM_KEYWORDS = [
    "estimate", "lineitem", "line_item", "item", "detail",
    "cost", "labor", "material", "scope",
]


def _parse_est_file(file_path: str) -> dict:
    """Parse a native WinEst .est file (OLE2/Compound Document format)."""
    try:
        import olefile
    except ImportError:
        return {
            "success": False,
            "format_detected": "est_native",
            "line_items": [],
            "warnings": [],
            "error": "olefile library not installed. Run: pip install olefile",
        }

    try:
        if not olefile.isOleFile(file_path):
            return {
                "success": False,
                "format_detected": "est_native",
                "line_items": [],
                "warnings": [],
                "error": _OLE_FALLBACK_ERROR,
            }

        ole = olefile.OleFileIO(file_path)
        streams = ole.listdir()
        stream_names = ["/".join(parts) for parts in streams]
        logger.info(f"WinEst .est streams ({len(streams)}): {stream_names[:20]}")

        warnings: list[str] = []
        line_items: list[dict] = []

        # Prioritise streams whose names suggest estimate data
        candidate_streams = [
            parts for parts in streams
            if any(kw in "/".join(parts).lower() for kw in _DATA_STREAM_KEYWORDS)
        ]
        if not candidate_streams:
            candidate_streams = streams[:8]
            warnings.append(
                f"No standard WinEst data streams found. "
                f"Streams present: {stream_names[:10]}. "
                "Results may be incomplete — consider exporting as .xlsx from WinEst."
            )

        parsed_any = False
        for stream_parts in candidate_streams:
            try:
                raw = ole.openstream(stream_parts).read()
                items = _extract_items_from_stream(raw, "/".join(stream_parts))
                if items:
                    line_items.extend(items)
                    parsed_any = True
            except Exception as exc:
                warnings.append(
                    f"Could not parse stream '{'/'.join(stream_parts)}': {exc}"
                )

        ole.close()

        if not parsed_any or not line_items:
            return {
                "success": False,
                "format_detected": "est_native",
                "line_items": [],
                "warnings": warnings,
                "error": _OLE_FALLBACK_ERROR,
            }

        return {
            "success": True,
            "format_detected": "est_native",
            "line_items": line_items,
            "warnings": warnings,
            "error": None,
        }

    except Exception as exc:
        logger.error(f"Error opening .est file '{file_path}': {exc}")
        return {
            "success": False,
            "format_detected": "est_native",
            "line_items": [],
            "warnings": [],
            "error": _OLE_FALLBACK_ERROR,
        }


def _extract_items_from_stream(data: bytes, stream_name: str) -> list[dict]:
    """Heuristically extract text records from a raw OLE2 stream.

    WinEst stores estimate data in a proprietary binary format that varies
    by version.  We scan for printable text using multiple encodings and
    surface each coherent string as a candidate line-item description.
    Numeric fields (quantity, cost, etc.) are left as None — downstream
    agents can fill them via gap analysis or takeoff.
    """
    items: list[dict] = []

    for encoding in ("utf-16-le", "latin-1", "utf-8"):
        try:
            text = data.decode(encoding, errors="replace")
            lines = [ln.strip() for ln in text.splitlines() if len(ln.strip()) > 3]
            # Keep lines that are mostly printable and not null-heavy
            printable = [
                ln for ln in lines
                if (sum(c.isprintable() for c in ln) / max(len(ln), 1)) > 0.70
                and not ln.startswith("\x00")
                and not ln.startswith("\ufffd" * 3)
            ]
            if len(printable) >= 3:
                for ln in printable[:200]:
                    clean = "".join(c for c in ln if c.isprintable()).strip()
                    if len(clean) > 4:
                        items.append({
                            "description":       clean,
                            "item_code":         None,
                            "quantity":          None,
                            "unit":              None,
                            "labor_hours":       None,
                            "labor_rate":        None,
                            "material_cost":     None,
                            "total":             None,
                            "csi_code":          None,
                            "wbs_code":          None,
                            "crew_size":         None,
                            "productivity_rate": None,
                            "source_stream":     stream_name,
                        })
                break  # stop trying encodings once we get usable text
        except Exception:
            continue

    return items


# ---------------------------------------------------------------------------
# Scenarios B & C — WinEst Excel exports (.xlsx / .xls)
# ---------------------------------------------------------------------------

def _parse_xlsx_file(file_path: str) -> dict:
    """Auto-detect WinEst format from column headers and parse accordingly."""
    try:
        import openpyxl
    except ImportError:
        return {
            "success": False,
            "format_detected": "unknown_xlsx",
            "line_items": [],
            "warnings": [],
            "error": "openpyxl library not installed.",
        }

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            wb.close()
            return {
                "success": False,
                "format_detected": "unknown_xlsx",
                "line_items": [],
                "warnings": ["Spreadsheet is empty or has no header row"],
                "error": None,
            }

        headers = {str(h).strip() for h in header_row if h is not None}

        if WINEST_FORMAT1_HEADERS.issubset(headers):
            result = _parse_format1(ws, header_row)
            result["format_detected"] = "xlsx_format1"
            wb.close()
            return result

        if WINEST_FORMAT2_HEADERS.issubset(headers):
            result = _parse_format2(ws, header_row)
            result["format_detected"] = "xlsx_format2"
            wb.close()
            return result

        wb.close()
        return {
            "success": False,
            "format_detected": "unknown_xlsx",
            "line_items": [],
            "warnings": [f"Headers found: {sorted(headers)}"],
            "error": None,
        }

    except Exception as exc:
        logger.error(f"Error parsing .xlsx file '{file_path}': {exc}")
        return {
            "success": False,
            "format_detected": "unknown_xlsx",
            "line_items": [],
            "warnings": [],
            "error": str(exc),
        }


# ---------------------------------------------------------------------------
# Column helpers
# ---------------------------------------------------------------------------

def _col_index(header_row: tuple, name: str) -> Optional[int]:
    """Return the 0-based index of *name* in *header_row*, or None."""
    for i, h in enumerate(header_row):
        if h is not None and str(h).strip() == name:
            return i
    return None


def _safe_float(value) -> Optional[float]:
    """Convert a spreadsheet cell value to float; return None on failure."""
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _cell(row: tuple, idx: Optional[int]) -> Optional[str]:
    """Return cell value as stripped string, or None if missing/empty."""
    if idx is None or row[idx] is None:
        return None
    v = str(row[idx]).strip()
    return v if v else None


# ---------------------------------------------------------------------------
# Format 1 parser
# ---------------------------------------------------------------------------

def _parse_format1(ws, header_row: tuple) -> dict:
    """Parse WinEst Format 1.

    Expected headers:
        Item | Description | Quantity | Unit | Labor Hours | Labor Rate | Material Cost | Total
    """
    col = {
        "item":          _col_index(header_row, "Item"),
        "description":   _col_index(header_row, "Description"),
        "quantity":      _col_index(header_row, "Quantity"),
        "unit":          _col_index(header_row, "Unit"),
        "labor_hours":   _col_index(header_row, "Labor Hours"),
        "labor_rate":    _col_index(header_row, "Labor Rate"),
        "material_cost": _col_index(header_row, "Material Cost"),
        "total":         _col_index(header_row, "Total"),
    }

    line_items: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = _cell(row, col["description"])
        if not desc:
            continue  # skip blank rows

        line_items.append({
            "description":       desc,
            "item_code":         _cell(row, col["item"]),
            "quantity":          _safe_float(row[col["quantity"]] if col["quantity"] is not None else None),
            "unit":              _cell(row, col["unit"]),
            "labor_hours":       _safe_float(row[col["labor_hours"]] if col["labor_hours"] is not None else None),
            "labor_rate":        _safe_float(row[col["labor_rate"]] if col["labor_rate"] is not None else None),
            "material_cost":     _safe_float(row[col["material_cost"]] if col["material_cost"] is not None else None),
            "total":             _safe_float(row[col["total"]] if col["total"] is not None else None),
            "csi_code":          None,
            "wbs_code":          None,
            "crew_size":         None,
            "productivity_rate": None,
        })

    return {
        "success": True,
        "line_items": line_items,
        "warnings": [],
        "error": None,
    }


# ---------------------------------------------------------------------------
# Format 2 parser
# ---------------------------------------------------------------------------

def _parse_format2(ws, header_row: tuple) -> dict:
    """Parse WinEst Format 2.

    Expected headers:
        WBS | CSI Code | Description | Qty | UOM | Crew Size | Productivity Rate | Hours | Cost
    """
    col = {
        "wbs":               _col_index(header_row, "WBS"),
        "csi_code":          _col_index(header_row, "CSI Code"),
        "description":       _col_index(header_row, "Description"),
        "qty":               _col_index(header_row, "Qty"),
        "uom":               _col_index(header_row, "UOM"),
        "crew_size":         _col_index(header_row, "Crew Size"),
        "productivity_rate": _col_index(header_row, "Productivity Rate"),
        "hours":             _col_index(header_row, "Hours"),
        "cost":              _col_index(header_row, "Cost"),
    }

    line_items: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = _cell(row, col["description"])
        if not desc:
            continue

        line_items.append({
            "description":       desc,
            "wbs_code":          _cell(row, col["wbs"]),
            "csi_code":          _cell(row, col["csi_code"]),
            "quantity":          _safe_float(row[col["qty"]] if col["qty"] is not None else None),
            "unit":              _cell(row, col["uom"]),
            "crew_size":         _safe_float(row[col["crew_size"]] if col["crew_size"] is not None else None),
            "productivity_rate": _safe_float(row[col["productivity_rate"]] if col["productivity_rate"] is not None else None),
            "labor_hours":       _safe_float(row[col["hours"]] if col["hours"] is not None else None),
            "total":             _safe_float(row[col["cost"]] if col["cost"] is not None else None),
            "labor_rate":        None,
            "material_cost":     None,
            "item_code":         None,
        })

    return {
        "success": True,
        "line_items": line_items,
        "warnings": [],
        "error": None,
    }
