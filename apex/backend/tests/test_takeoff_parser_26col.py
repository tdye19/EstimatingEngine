"""Tests for the 26-column CCI Civil Est Report takeoff parser.

Covers the column-mapping and skip-rule fix from spec 18.3.3.3:
activity reads from col E (not C), production_rate from col H (not I),
labor_cost_per_unit from col M (not K), material_cost_per_unit from col N
(not L), and group-marker / rollup rows are filtered.
"""

import pytest
from openpyxl import Workbook

from apex.backend.services.takeoff_parser.parser import (
    detect_takeoff_format,
    parse_26col_takeoff,
)


@pytest.fixture
def aws_sample_26col_xlsx(tmp_path):
    """Mimics the real AWS WinEst .xlsx layout: 27 columns (A–AA), header
    at xlsx row 4, data starting row 5. Includes group-rollup rows,
    WBS-rollup rows, two real line items, and two skip-worthy rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(row=1, column=1, value="_CCI Civil Est Report")

    headers = [
        "Alternate Name",                         # A
        "<Estimate.Wbs8TypeName> Name",           # B
        "Alternate Name",                         # C
        "Note",                                   # D
        "Item Description",                       # E
        "Takeoff Quantity - Adjusted",            # F
        "Unit",                                   # G
        "Labor Productivity - Adjusted (TradeHours)",  # H
        "Labor Prod Unit",                        # I
        "Labor Mix",                              # J
        "Labor Mix 1 Value",                      # K
        "Labor Hours (TradeHours)",               # L
        "Labor Unit Price",                       # M
        "Mat Unit Price",                         # N
        "Equip Prod",                             # O
        "Equip Prod Unit",                        # P
        "Equip Mix",                              # Q
        "Equip Unit Price",                       # R
        "Subs Unit Price",                        # S
        "Labor Total",                            # T
        "Mat Total",                              # U
        "Equip Total",                            # V
        "Subs Total",                             # W
        "Other Total",                            # X
        "col Y",                                  # Y
        "Total Unit Price",                       # Z
        "Grand Total",                            # AA
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)

    # Row 5 — full-estimate group rollup (A, C, E all "--- Base Estimate ---",
    # no qty, grand total in AA). Should be SKIPPED.
    ws.cell(row=5, column=1, value="--- Base Estimate ---")
    ws.cell(row=5, column=3, value="--- Base Estimate ---")
    ws.cell(row=5, column=5, value="--- Base Estimate ---")
    ws.cell(row=5, column=27, value=1_000_000)

    # Row 6 — WBS-level rollup (description present but no F quantity).
    # Should be SKIPPED by the "quantity is None or <= 0" rule.
    ws.cell(row=6, column=2, value="General Conditions")
    ws.cell(row=6, column=5, value="General Conditions")
    ws.cell(row=6, column=12, value=100)
    ws.cell(row=6, column=27, value=50_000)

    # Row 7 — real line item (Senior Project Manager).
    ws.cell(row=7, column=1, value="--- Base Estimate ---")
    ws.cell(row=7, column=2, value="General Conditions")
    ws.cell(row=7, column=3, value="--- Base Estimate ---")
    ws.cell(row=7, column=5, value="Senior Project Manager")
    ws.cell(row=7, column=6, value=76)
    ws.cell(row=7, column=7, value="week")
    ws.cell(row=7, column=8, value=0.02)
    ws.cell(row=7, column=9, value="week/hour")
    ws.cell(row=7, column=13, value=4250)
    ws.cell(row=7, column=27, value=323_000)

    # Row 8 — real line item (Curb/Curb & Gutter).
    ws.cell(row=8, column=2, value="Curbs")
    ws.cell(row=8, column=5, value="Curb/Curb & Gutter Concrete 4000psi")
    ws.cell(row=8, column=6, value=354)
    ws.cell(row=8, column=7, value="cuyd")
    ws.cell(row=8, column=14, value=180)
    ws.cell(row=8, column=26, value=180)
    ws.cell(row=8, column=27, value=63_714)

    # Row 9 — qty but no description. Should be SKIPPED.
    ws.cell(row=9, column=6, value=999)

    # Row 10 — description but non-numeric qty. Should be SKIPPED.
    ws.cell(row=10, column=5, value="Subtotal - Civil")
    ws.cell(row=10, column=6, value="TBD")
    ws.cell(row=10, column=7, value="week")

    path = tmp_path / "aws_sample_26col.xlsx"
    wb.save(path)
    return str(path)


class TestParse26ColTakeoff:
    def test_parses_exactly_two_real_items(self, aws_sample_26col_xlsx):
        items = parse_26col_takeoff(aws_sample_26col_xlsx)
        assert len(items) == 2

    def test_first_item_senior_pm(self, aws_sample_26col_xlsx):
        items = parse_26col_takeoff(aws_sample_26col_xlsx)
        pm = items[0]
        assert pm.activity == "Senior Project Manager"
        assert pm.quantity == 76
        assert pm.unit == "week"
        assert pm.wbs_area == "General Conditions"
        assert pm.production_rate == 0.02
        assert pm.labor_cost_per_unit == 4250

    def test_second_item_curb(self, aws_sample_26col_xlsx):
        items = parse_26col_takeoff(aws_sample_26col_xlsx)
        curb = items[1]
        assert curb.activity == "Curb/Curb & Gutter Concrete 4000psi"
        assert curb.quantity == 354
        assert curb.unit == "cuyd"
        assert curb.wbs_area == "Curbs"
        assert curb.material_cost_per_unit == 180

    def test_no_group_marker_activity(self, aws_sample_26col_xlsx):
        """Regression guard for the col-C bug: no item should have activity
        equal to '--- Base Estimate ---'."""
        items = parse_26col_takeoff(aws_sample_26col_xlsx)
        assert all(item.activity != "--- Base Estimate ---" for item in items)
        assert all(not item.activity.startswith("---") for item in items)

    def test_crew_is_none_in_26col(self, aws_sample_26col_xlsx):
        """26-col format has no crew column — parser should emit None."""
        items = parse_26col_takeoff(aws_sample_26col_xlsx)
        assert all(item.crew is None for item in items)

    def test_row_numbers_contiguous(self, aws_sample_26col_xlsx):
        """Skipped rows must not leave gaps in row_number sequence."""
        items = parse_26col_takeoff(aws_sample_26col_xlsx)
        assert [i.row_number for i in items] == [1, 2]


class TestDetect26ColFormat:
    def test_a1_signature_detected(self, aws_sample_26col_xlsx):
        """Regression guard: A1 cell starting with '_CCI Civil Est Report'
        must still route to the 26col parser path."""
        assert detect_takeoff_format(aws_sample_26col_xlsx) == "26col"
