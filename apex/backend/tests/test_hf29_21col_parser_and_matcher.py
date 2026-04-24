"""HF-29 regression tests — 21-col parser column remap + matcher prefers file rates.

Project 21 returned proposal_form.base_bid.total = 0 because:
  - The 21-col parser was reading 2 columns to the LEFT of where it
    should have been (col C as activity, col D as quantity, etc.), so
    100% of "items" were the rollup marker "--- Base Estimate ---".
  - It also didn't read price columns at all (cols L/M).
  - Even when items DID have parser-extracted rates, the matcher
    silently overwrote them with PB historical averages.

HF-29 fixes:
  Edit 1 — parse_21col_takeoff column remap (D→wbs_area, E→activity,
           F→quantity, G→unit, H→crew, I→production_rate,
           L→labor_cost_per_unit, M→material_cost_per_unit) +
           "---" and qty>0 filters mirroring the 26-col path.
  Edit 3 — RateMatchingEngine: prefer item.labor_cost_per_unit /
           material_cost_per_unit (estimator's file rate) over PB avg;
           PB avg is fallback only when parser produced None.
"""

from __future__ import annotations

import uuid
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from apex.backend.agents.agent_4_takeoff import run_takeoff_agent
from apex.backend.agents.proposal_form_builder import build_proposal_form
from apex.backend.models.document import Document
from apex.backend.models.line_item_wc_attribution import LineItemWCAttribution
from apex.backend.models.project import Project
from apex.backend.models.takeoff_v2 import TakeoffItemV2
from apex.backend.models.work_category import WorkCategory
from apex.backend.services.library.productivity_brain.models import (
    PBLineItem,
    PBProject,
)
from apex.backend.services.takeoff_parser.parser import (
    parse_21col_takeoff,
    parse_26col_takeoff,
)

KCCU_FIXTURE = (
    Path(__file__).parent / "fixtures" / "takeoff" / "kccu_concrete_winest_21col.xlsx"
)


# ---------------------------------------------------------------------------
# Test 1 — KCCU 21-col parser pulls real rows + correct rates
# ---------------------------------------------------------------------------


def test_kccu_concrete_winest_reads_rates_correctly():
    """First 5 real data rows from the KCCU Concrete winest fixture must
    match the known-good values verified directly from the xlsx (xlsx rows
    7-11 after rollup + section-header filtering)."""
    items = parse_21col_takeoff(str(KCCU_FIXTURE))

    # Total real-data rows: 121 (vs. pre-HF-29 parser's 114 garbage items).
    assert len(items) >= 100, f"expected ~121 real items, got {len(items)}"

    # Known-good values, captured from openpyxl read of the actual xlsx.
    expected = [
        # (activity, qty, unit, production_rate, labor_cost_per_unit, material_cost_per_unit)
        ("Project Manager", 5, "week", 0.13, 640, None),
        ("Project Superintendent", 5, "week", 0.03, 3200, None),
        ("Project Engineer", 5, "week", 0.06, 1040, None),
        ("Field Layout Engineering/Survey", 2, "week", 0.05, 1500, None),
        ("Trade Travel", 5000, "lsum", None, None, 0.5),
    ]

    for i, (act, qty, unit, rate, lab, mat) in enumerate(expected):
        item = items[i]
        assert item.activity == act, f"item {i}: activity {item.activity!r} != {act!r}"
        assert item.quantity == qty, f"item {i}: qty {item.quantity} != {qty}"
        assert item.unit == unit, f"item {i}: unit {item.unit!r} != {unit!r}"
        assert item.production_rate == rate, (
            f"item {i}: production_rate {item.production_rate} != {rate}"
        )
        assert item.labor_cost_per_unit == lab, (
            f"item {i}: labor_cost_per_unit {item.labor_cost_per_unit} != {lab}"
        )
        assert item.material_cost_per_unit == mat, (
            f"item {i}: material_cost_per_unit {item.material_cost_per_unit} != {mat}"
        )

    # All items have wbs_area = "General Conditions" (first 5 are all in that
    # WBS group); pinning the column-D mapping.
    for i in range(5):
        assert items[i].wbs_area == "General Conditions"


# ---------------------------------------------------------------------------
# Test 2 — Rollup row filter invariant (Tucker Observation 1)
# ---------------------------------------------------------------------------


def test_21col_rollup_row_filtered(tmp_path):
    """Synthetic fixture: only one data row, preceded by a rollup row.
    The "---" filter must catch the rollup, and the qty>0 filter must
    catch a section-header row with no quantity. Pins the filter
    invariants so a future refactor can't quietly drop them."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="_CCI Estimate Report")
    # Header at xlsx row 4 (cols D=wbs, E=activity, F=qty, G=unit, etc.)
    ws.cell(row=4, column=4, value="WBS8 Name")
    ws.cell(row=4, column=5, value="Item Description")
    ws.cell(row=4, column=6, value="Takeoff Quantity - Adjusted")
    ws.cell(row=4, column=7, value="Unit")
    ws.cell(row=4, column=12, value="Labor Unit Price")
    ws.cell(row=4, column=13, value="Mat Unit Price")

    # Row 5 — full-estimate rollup. activity matches "---" filter.
    ws.cell(row=5, column=5, value="--- Base Estimate ---")

    # Row 6 — section header. activity present but qty is None.
    ws.cell(row=6, column=4, value="General Conditions")
    ws.cell(row=6, column=5, value="General Conditions")

    # Row 7 — real data. Should be the ONLY row that survives.
    ws.cell(row=7, column=4, value="General Conditions")
    ws.cell(row=7, column=5, value="Project Manager")
    ws.cell(row=7, column=6, value=5)
    ws.cell(row=7, column=7, value="week")
    ws.cell(row=7, column=12, value=640)

    path = tmp_path / "rollup_filter_pin.xlsx"
    wb.save(path)

    items = parse_21col_takeoff(str(path))
    assert len(items) == 1
    assert items[0].activity == "Project Manager"
    assert items[0].labor_cost_per_unit == 640


# ---------------------------------------------------------------------------
# Test 3 — AWS 26-col parser unchanged (HF-29 must not touch this path)
# ---------------------------------------------------------------------------


def _build_aws_26col_fixture(tmp_path: Path) -> Path:
    """Minimal 26-col fixture — mirrors test_takeoff_parser_26col.py's
    aws_sample_26col_xlsx fixture but is local to this file so the test
    can run independently."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="_CCI Civil Est Report")

    # Header at xlsx row 4. We only set the cells the parser actually reads.
    ws.cell(row=4, column=2, value="WBS Name")
    ws.cell(row=4, column=5, value="Item Description")
    ws.cell(row=4, column=6, value="Takeoff Quantity - Adjusted")
    ws.cell(row=4, column=7, value="Unit")
    ws.cell(row=4, column=8, value="Labor Productivity")
    ws.cell(row=4, column=13, value="Labor Unit Price")
    ws.cell(row=4, column=14, value="Mat Unit Price")

    # Row 5 rollup → must be skipped.
    ws.cell(row=5, column=5, value="--- Base Estimate ---")
    # Row 6 section header (no qty) → must be skipped.
    ws.cell(row=6, column=5, value="General Conditions")
    # Row 7 real item.
    ws.cell(row=7, column=2, value="General Conditions")
    ws.cell(row=7, column=5, value="Senior Project Manager")
    ws.cell(row=7, column=6, value=76)
    ws.cell(row=7, column=7, value="week")
    ws.cell(row=7, column=8, value=0.02)
    ws.cell(row=7, column=13, value=4250)
    # Row 8 second real item.
    ws.cell(row=8, column=2, value="Curbs")
    ws.cell(row=8, column=5, value="Curb/Curb & Gutter Concrete 4000psi")
    ws.cell(row=8, column=6, value=354)
    ws.cell(row=8, column=7, value="cuyd")
    ws.cell(row=8, column=14, value=180)

    path = tmp_path / "aws_26col_regression.xlsx"
    wb.save(path)
    return path


def test_aws_winest_rates_unchanged_post_hf29(tmp_path):
    """Sprint 18.3.3.3's AWS 26-col path is untouched by HF-29's parser
    edits. Pin the parser-side rates so a future change to the 21-col
    code can't silently regress this format."""
    fixture = _build_aws_26col_fixture(tmp_path)
    items = parse_26col_takeoff(str(fixture))

    assert len(items) == 2
    assert items[0].activity == "Senior Project Manager"
    assert items[0].labor_cost_per_unit == 4250
    assert items[0].material_cost_per_unit is None
    assert items[1].activity == "Curb/Curb & Gutter Concrete 4000psi"
    assert items[1].material_cost_per_unit == 180
    assert items[1].labor_cost_per_unit is None


# ---------------------------------------------------------------------------
# Test 4 — Edit 3 semantic pin via Agent 4 persistence (Tucker Observation 4)
# ---------------------------------------------------------------------------


def test_aws_winest_persists_estimator_rates_not_pb_avg(
    db_session: Session, tmp_path
):
    """End-to-end Agent 4 path: when the file's parser-extracted rate
    differs from the PB historical average for the same activity, the
    persisted TakeoffItemV2.labor_cost_per_unit must equal the FILE rate
    (estimator's bid), not the PB avg. Edit 3 happens in the matcher;
    this pins behavior at the persistence layer (where the proposal_form
    actually reads)."""
    suffix = uuid.uuid4().hex[:8]
    project = Project(
        name=f"HF29 e2e {suffix}",
        project_number=f"HF29-E2E-{suffix}",
        project_type="commercial",
    )
    db_session.add(project)
    db_session.commit()
    db_session.refresh(project)

    # Build an AWS 26-col fixture with one real item — file labor_cost = 4250.
    fixture = _build_aws_26col_fixture(tmp_path)
    doc = Document(
        project_id=project.id,
        filename="aws_test.xlsx",
        file_path=str(fixture),
        file_type="xlsx",
        classification="winest",
        processing_status="completed",
    )
    db_session.add(doc)
    db_session.commit()

    # Seed PB with a matching activity but a DIFFERENT avg labor cost (50.00)
    # so we can prove which value gets persisted.
    pb_proj = PBProject(
        name="seed-project",
        source_file="seed.xlsx",
        file_hash=uuid.uuid4().hex,
        format_type="26col_civil",
    )
    db_session.add(pb_proj)
    db_session.commit()
    db_session.refresh(pb_proj)
    pb_row = PBLineItem(
        project_id=pb_proj.id,
        source_project="seed-project",
        activity="Senior Project Manager",
        unit="week",
        production_rate=0.02,
        labor_cost_per_unit=50.00,    # ← deliberately different from file's 4250
        material_cost_per_unit=10.00,
    )
    db_session.add(pb_row)
    db_session.commit()

    result = run_takeoff_agent(db_session, project.id)
    assert result["takeoff_items_parsed"] == 2

    rows = (
        db_session.query(TakeoffItemV2)
        .filter_by(project_id=project.id)
        .order_by(TakeoffItemV2.row_number)
        .all()
    )
    pm_row = next(r for r in rows if r.activity == "Senior Project Manager")
    # The file rate (4250) wins. PB avg (50.00) is not what's persisted.
    assert pm_row.labor_cost_per_unit == 4250, (
        f"persisted labor_cost_per_unit={pm_row.labor_cost_per_unit} — "
        "Edit 3 broken: matcher overwrote estimator's file rate with PB avg"
    )
    # historical_avg_rate field still reflects PB data (separate column).
    assert pm_row.historical_avg_rate == 0.02


# ---------------------------------------------------------------------------
# Test 5 — proposal_form base_bid is non-zero with real rates
# ---------------------------------------------------------------------------


def test_proposal_form_base_bid_nonzero_with_real_rates(db_session: Session):
    """End-to-end proposal_form arithmetic: TakeoffItemV2 rows seeded with
    non-zero rates produce base_bid.total > 0. Pins the dollar-flow from
    parser → matcher → DB → proposal_form against silent zeroing."""
    suffix = uuid.uuid4().hex[:8]
    project = Project(
        name=f"HF29 pf {suffix}",
        project_number=f"HF29-PF-{suffix}",
        project_type="commercial",
    )
    db_session.add(project)
    db_session.commit()
    db_session.refresh(project)

    wc = WorkCategory(
        project_id=project.id,
        wc_number="WC 02",
        title="Concrete",
        work_included_items=[],
        specific_notes=[],
        related_work_by_others=[],
        add_alternates=[],
        allowances=[],
        unit_prices=[],
        referenced_spec_sections=[],
    )
    db_session.add(wc)
    db_session.commit()
    db_session.refresh(wc)

    # Two seeded TakeoffItemV2 rows with KCCU-style rates.
    items = [
        TakeoffItemV2(
            project_id=project.id, row_number=1, activity="Project Manager",
            quantity=5, unit="week", labor_cost_per_unit=640, material_cost_per_unit=None,
        ),
        TakeoffItemV2(
            project_id=project.id, row_number=2, activity="Trade Travel",
            quantity=5000, unit="lsum", labor_cost_per_unit=None, material_cost_per_unit=0.5,
        ),
    ]
    for t in items:
        db_session.add(t)
    db_session.commit()
    for t in items:
        db_session.refresh(t)
        db_session.add(
            LineItemWCAttribution(
                project_id=project.id, takeoff_item_id=t.id,
                work_category_id=wc.id, match_tier="csi_exact",
                confidence=1.0, source="rule", rationale="seed",
            )
        )
    db_session.commit()

    pf = build_proposal_form(db_session, project.id)
    assert pf is not None
    # 5 * 640 + 5000 * 0.5 = 3200 + 2500 = 5700.00
    assert pf["base_bid"]["total"] == 5700.0
    wc_row = pf["base_bid"]["by_work_category"][0]
    assert wc_row["wc_number"] == "WC 02"
    assert wc_row["labor_cost"] == 3200.0
    assert wc_row["material_cost"] == 2500.0
    assert wc_row["subtotal"] == 5700.0
