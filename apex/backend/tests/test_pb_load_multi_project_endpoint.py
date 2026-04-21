"""DATA-1.2 — POST /api/library/productivity-brain/load-multi-project tests."""

from pathlib import Path

import pytest
from openpyxl import Workbook

from apex.backend.services.library.productivity_brain.models import PBLineItem, PBProject

URL = "/api/library/productivity-brain/load-multi-project"
_FIXTURE = Path(__file__).parent / "fixtures" / "pb" / "CityGate_Master_Productivity_Rates.xlsx"


@pytest.fixture(autouse=True)
def _clean_pb_tables(db_session):
    db_session.query(PBLineItem).delete()
    db_session.query(PBProject).delete()
    db_session.commit()
    yield


@pytest.fixture
def citygate_bytes():
    if not _FIXTURE.exists():
        pytest.skip(f"CityGate fixture not present at {_FIXTURE}")
    return _FIXTURE.read_bytes()


@pytest.fixture
def garbage_xlsx_bytes(tmp_path):
    """A valid xlsx that is NOT multi-project rates — used to exercise 400."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Totally unrelated report")
    ws.cell(row=2, column=1, value="No useful headers here")
    path = tmp_path / "garbage.xlsx"
    wb.save(path)
    return path.read_bytes()


class TestLoadMultiProjectEndpoint:
    def test_unauthenticated_returns_401(self, client):
        res = client.post(URL, files={"file": ("x.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert res.status_code == 401

    def test_non_admin_returns_403(self, client, auth_headers, citygate_bytes):
        res = client.post(
            URL,
            headers=auth_headers,
            files={"file": ("CityGate.xlsx", citygate_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 403

    def test_happy_path(self, client, admin_headers, citygate_bytes, db_session):
        res = client.post(
            URL,
            headers=admin_headers,
            files={"file": ("CityGate.xlsx", citygate_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"metadata_json": '{"region":"CCI Outstate","customer":"Consumers Energy"}'},
        )
        assert res.status_code == 200, res.text
        body = res.json()
        assert body["success"] is True
        data = body["data"]
        assert data["projects_upserted_new"] == 4
        assert data["projects_upserted_existing"] == 0
        assert data["line_items_inserted"] > 0
        assert len(data["pb_project_ids"]) == 4
        assert data["metadata_applied"] == {"region": "CCI Outstate", "customer": "Consumers Energy"}
        # Projects really landed
        assert db_session.query(PBProject).count() == 4

    def test_reupload_same_file_returns_409(self, client, admin_headers, citygate_bytes):
        first = client.post(
            URL,
            headers=admin_headers,
            files={"file": ("CityGate.xlsx", citygate_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert first.status_code == 200

        second = client.post(
            URL,
            headers=admin_headers,
            files={"file": ("CityGate.xlsx", citygate_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert second.status_code == 409
        body = second.json()
        assert body["success"] is False
        assert body["error"] == "duplicate_file"
        assert len(body["data"]["existing_project_ids"]) == 4
        assert len(body["data"]["existing_project_names"]) == 4

    def test_garbage_xlsx_returns_400(self, client, admin_headers, garbage_xlsx_bytes):
        res = client.post(
            URL,
            headers=admin_headers,
            files={"file": ("garbage.xlsx", garbage_xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 400
        assert "multi-project rates" in res.json()["detail"].lower()

    def test_non_xlsx_extension_returns_400(self, client, admin_headers):
        res = client.post(
            URL,
            headers=admin_headers,
            files={"file": ("x.csv", b"a,b,c\n1,2,3\n", "text/csv")},
        )
        assert res.status_code == 400

    def test_malformed_metadata_json_returns_400(self, client, admin_headers, citygate_bytes):
        res = client.post(
            URL,
            headers=admin_headers,
            files={"file": ("CityGate.xlsx", citygate_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"metadata_json": "{not valid json"},
        )
        assert res.status_code == 400

    def test_metadata_json_non_object_returns_400(self, client, admin_headers, citygate_bytes):
        res = client.post(
            URL,
            headers=admin_headers,
            files={"file": ("CityGate.xlsx", citygate_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"metadata_json": '["region","CCI Outstate"]'},
        )
        assert res.status_code == 400
