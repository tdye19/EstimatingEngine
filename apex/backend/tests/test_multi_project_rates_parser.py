"""DATA-1.1 — multi-project rates parser + loader tests.

Real-file tests run against the CityGate fixture at
  apex/backend/tests/fixtures/pb/CityGate_Master_Productivity_Rates.xlsx
That path is gitignored; the marker below skips when absent rather than
falling back to a synthetic fixture.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from apex.backend.services.library.productivity_brain.models import PBLineItem, PBProject
from apex.backend.services.library.productivity_brain.parsers import MultiProjectRatesParser
from apex.backend.services.library.productivity_brain.service import ProductivityBrainService

_FIXTURE = Path(__file__).parent / "fixtures" / "pb" / "CityGate_Master_Productivity_Rates.xlsx"

_CITYGATE_SECTIONS = {
    "000 — General Conditions",
    "010 — M&R Building",
    "015 — Site Concrete",
    "015 — Site Concrete (Storm)",
    "030 — Filter / Separator",
    "040 — Heater",
    "050 — RTU Rack",
    "060 — Light Pole",
    "080 — Misc. Concrete",
}


@pytest.fixture(autouse=True)
def _clean_pb_tables(db_session):
    db_session.query(PBLineItem).delete()
    db_session.query(PBProject).delete()
    db_session.commit()
    yield


@pytest.fixture
def citygate_path():
    if not _FIXTURE.exists():
        pytest.skip(f"CityGate fixture not present at {_FIXTURE}")
    return str(_FIXTURE)


@pytest.fixture
def cci_26col_xlsx(tmp_path):
    """Synthetic 26-column CCI Civil Est Report — used only as a negative
    case for detect(). Mirrors the layout flags the existing averaged-rates
    parser looks at so the structure is plausibly recognisable."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # A1 is what the legacy detect() keys off for 26col_civil
    ws.cell(row=1, column=1, value="12345_CCI Civil Est Report")
    ws.cell(row=2, column=1, value="Subtitle")
    # Pad out ~26 columns with arbitrary headers; no multi-project header row
    headers = [
        "Item", "WBS", "Code", "Description", "Qty", "Unit", "Prod", "Prod Unit", "Crew",
        "Hours", "Labor Hrs", "Labor UP", "Mat UP", "c13", "c14", "c15", "c16",
        "Equip UP", "Subs UP", "Labor Total", "Mat Total", "Equip Total", "Subs Total",
        "c23", "c24", "Grand Total",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    path = tmp_path / "fake_cci_26col.xlsx"
    wb.save(path)
    return str(path)


class TestDetect:
    def test_true_on_citygate(self, citygate_path):
        assert MultiProjectRatesParser.detect(citygate_path) is True

    def test_false_on_cci_26col(self, cci_26col_xlsx):
        assert MultiProjectRatesParser.detect(cci_26col_xlsx) is False

    def test_false_on_nonexistent_path(self, tmp_path):
        assert MultiProjectRatesParser.detect(str(tmp_path / "nope.xlsx")) is False


class TestParser:
    def test_produces_4_projects(self, citygate_path):
        result = MultiProjectRatesParser().parse(citygate_path)
        assert len(result.parsed_projects) == 4
        assert [p.source_project for p in result.parsed_projects] == [
            "Flint", "Bancroft", "Hanover", "Highland",
        ]
        assert [p.project_name for p in result.parsed_projects] == [
            "CCI CityGate Flint",
            "CCI CityGate Bancroft",
            "CCI CityGate Hanover",
            "CCI CityGate Highland",
        ]

    def test_per_project_line_item_count_bounds(self, citygate_path):
        result = MultiProjectRatesParser().parse(citygate_path)
        for proj in result.parsed_projects:
            assert 1 <= len(proj.line_items) <= 75, (
                f"{proj.source_project} emitted {len(proj.line_items)} items"
            )

    def test_all_production_rates_are_floats(self, citygate_path):
        result = MultiProjectRatesParser().parse(citygate_path)
        for proj in result.parsed_projects:
            for li in proj.line_items:
                assert isinstance(li.production_rate, float)

    def test_sentinels_skipped_not_coerced(self, citygate_path):
        """'Pilaster Forms - Wood' has '—' for three of four projects in the
        source file → only one project should have that activity."""
        result = MultiProjectRatesParser().parse(citygate_path)
        by_project_acts = {
            p.source_project: {li.activity_description for li in p.line_items}
            for p in result.parsed_projects
        }
        projects_with_pilaster = [
            name for name, acts in by_project_acts.items()
            if "Pilaster Forms - Wood" in acts
        ]
        assert len(projects_with_pilaster) == 1, (
            f"expected exactly 1 project to have 'Pilaster Forms - Wood'; "
            f"got {projects_with_pilaster}"
        )

    def test_section_header_rows_not_emitted_as_activities(self, citygate_path):
        result = MultiProjectRatesParser().parse(citygate_path)
        for proj in result.parsed_projects:
            for li in proj.line_items:
                assert li.activity_description not in _CITYGATE_SECTIONS

    def test_legend_rows_not_emitted(self, citygate_path):
        result = MultiProjectRatesParser().parse(citygate_path)
        for proj in result.parsed_projects:
            acts = {li.activity_description for li in proj.line_items}
            assert not any(a.startswith("Green =") for a in acts)
            assert not any(a.startswith("Spread =") for a in acts)

    def test_wbs_area_is_section_header(self, citygate_path):
        result = MultiProjectRatesParser().parse(citygate_path)
        for proj in result.parsed_projects:
            for li in proj.line_items:
                assert li.wbs_area in _CITYGATE_SECTIONS

    def test_field_layout_uniform_across_projects(self, citygate_path):
        """Source file has rate 0.04 for all 4 projects on row 5."""
        result = MultiProjectRatesParser().parse(citygate_path)
        hits = []
        for proj in result.parsed_projects:
            for li in proj.line_items:
                if li.activity_description == "Field Layout Engineering/Survey":
                    hits.append((proj.source_project, li.production_rate, li.unit))
        assert len(hits) == 4
        assert all(rate == 0.04 for _, rate, _ in hits)
        assert all(unit == "week" for _, _, unit in hits)

    def test_parse_raises_on_unrecognised_file(self, cci_26col_xlsx):
        with pytest.raises(ValueError, match="multi-project rates"):
            MultiProjectRatesParser().parse(cci_26col_xlsx)

    def test_metadata_override_flows_into_parsed_project(self, citygate_path):
        result = MultiProjectRatesParser().parse(
            citygate_path,
            metadata_overrides={"region": "CCI Outstate", "customer": "Consumers Energy"},
        )
        for proj in result.parsed_projects:
            assert proj.metadata["region"] == "CCI Outstate"
            assert proj.metadata["customer"] == "Consumers Energy"
            assert proj.metadata["file_source"] == "CityGate_Master_Productivity_Rates.xlsx"


class TestLoader:
    def test_initial_load_counts(self, db_session, citygate_path):
        svc = ProductivityBrainService(db_session)
        r = svc.load_multi_project_file(db_session, citygate_path)
        assert r.projects_upserted_new == 4
        assert r.projects_upserted_existing == 0
        assert r.line_items_inserted > 0
        assert r.line_items_updated == 0
        assert len(r.pb_project_ids) == 4
        assert db_session.query(PBProject).count() == 4

    def test_upsert_idempotent_on_rerun(self, db_session, citygate_path):
        svc = ProductivityBrainService(db_session)
        svc.load_multi_project_file(db_session, citygate_path)
        proj_count_1 = db_session.query(PBProject).count()
        li_count_1 = db_session.query(PBLineItem).count()

        r = svc.load_multi_project_file(db_session, citygate_path)
        assert db_session.query(PBProject).count() == proj_count_1
        assert db_session.query(PBLineItem).count() == li_count_1
        assert r.projects_upserted_new == 0
        assert r.projects_upserted_existing == 4
        assert r.line_items_updated == li_count_1
        assert r.line_items_inserted == 0

    def test_upsert_updates_mutated_rate(self, db_session, citygate_path):
        svc = ProductivityBrainService(db_session)
        svc.load_multi_project_file(db_session, citygate_path)

        victim = db_session.query(PBLineItem).first()
        original_rate = victim.production_rate
        victim.production_rate = -999.0
        db_session.commit()

        svc.load_multi_project_file(db_session, citygate_path)
        db_session.refresh(victim)
        assert victim.production_rate == original_rate

    def test_each_project_has_source_project_tag(self, db_session, citygate_path):
        svc = ProductivityBrainService(db_session)
        svc.load_multi_project_file(db_session, citygate_path)

        for name in ("Flint", "Bancroft", "Hanover", "Highland"):
            proj = db_session.query(PBProject).filter_by(name=f"CCI CityGate {name}").one()
            items = db_session.query(PBLineItem).filter_by(project_id=proj.id).all()
            assert len(items) > 0
            assert {li.source_project for li in items} == {name}

    def test_metadata_applied_returned_in_load_result(self, db_session, citygate_path):
        svc = ProductivityBrainService(db_session)
        r = svc.load_multi_project_file(
            db_session,
            citygate_path,
            metadata_overrides={"region": "CCI Outstate"},
        )
        assert r.metadata_applied == {"region": "CCI Outstate"}

    def test_load_rejects_unrecognised_file(self, db_session, cci_26col_xlsx):
        svc = ProductivityBrainService(db_session)
        with pytest.raises(ValueError, match="multi-project rates"):
            svc.load_multi_project_file(db_session, cci_26col_xlsx)
