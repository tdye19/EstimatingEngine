"""Sprint 18.4.3 — Proposal Form Excel export tests.

Covers the rendering service (synthetic JSON fixtures → xlsx bytes) and the
HTTP endpoint (auth/404/content-type). All tests use deterministic synthetic
payloads; no pipeline runs are required.
"""

from __future__ import annotations

import json
import uuid
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from apex.backend.models.intelligence_report import IntelligenceReportModel
from apex.backend.models.project import Project
from apex.backend.services.proposal_form_excel_service import (
    WARN_ALTS_NO_PRICE,
    WARN_FILL,
    WARN_UNATTRIBUTED,
    build_export_filename,
    render_proposal_form_xlsx,
)

EXPECTED_SHEET_NAMES = [
    "Summary",
    "Base Bid",
    "Alternates",
    "Allowances",
    "Unit Prices",
    "Inclusions & Exclusions",
    "Warnings & QC",
]


# ---------------------------------------------------------------------------
# Fixture factory
# ---------------------------------------------------------------------------
def make_proposal_form_json(**overrides) -> dict:
    """Return a valid proposal_form_json dict with sensible defaults.

    Keys in ``overrides`` replace defaults wholesale. ``warnings`` is a list,
    ``base_bid`` / ``alternates`` / ``allowances`` / ``unit_prices`` /
    ``breakout_notes`` follow the 18.4.2 ProposalForm contract.
    """
    base: dict = {
        "project_id": 1,
        "project_name": "Test Project",
        "generated_at": "2026-04-25T00:00:00+00:00",
        "base_bid": {
            "total": 3000.0,
            "by_work_category": [
                {
                    "wc_number": "10",
                    "wc_title": "Earthwork",
                    "line_items_count": 2,
                    "labor_cost": 500.0,
                    "material_cost": 500.0,
                    "subtotal": 1000.0,
                    "attribution_confidence_avg": 0.9,
                },
                {
                    "wc_number": "20",
                    "wc_title": "Concrete",
                    "line_items_count": 1,
                    "labor_cost": 1000.0,
                    "material_cost": 1000.0,
                    "subtotal": 2000.0,
                    "attribution_confidence_avg": 0.85,
                },
            ],
            "unattributed": None,
        },
        "alternates": [
            {
                "wc_number": "10",
                "description": "Swap aggregate base",
                "price_type": "add",
                "amount": 2500.0,
                "source": "work_category.add_alternates",
            }
        ],
        "allowances": [
            {
                "wc_number": "20",
                "description": "Contingency",
                "amount": 5000.0,
                "source": "work_category.allowances",
            }
        ],
        "unit_prices": [
            {
                "wc_number": "10",
                "description": "Extra excavation",
                "unit": "CY",
                "rate": 45.0,
                "source": "work_category.unit_prices",
            }
        ],
        "breakout_notes": [
            {
                "wc_number": "20",
                "description": "Breakout cost on proposal form — formwork",
                "source": "work_category.specific_notes",
            }
        ],
        "warnings": [],
    }
    base.update(overrides)
    return base


def _project(db: Session, tag: str = "x", name: str | None = None) -> Project:
    suffix = uuid.uuid4().hex[:8]
    p = Project(
        name=name or f"S1843 {tag}",
        project_number=f"S1843-{tag}-{suffix}",
        project_type="commercial",
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


def _write_report(db: Session, project: Project, payload: dict | None) -> IntelligenceReportModel:
    report = IntelligenceReportModel(
        project_id=project.id,
        version=1,
        proposal_form_json=json.dumps(payload) if payload is not None else None,
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return report


def _load(xlsx_bytes: bytes):
    return load_workbook(BytesIO(xlsx_bytes))


# ---------------------------------------------------------------------------
# Service tests
# ---------------------------------------------------------------------------
def test_renders_all_seven_sheets(db_session: Session):
    project = _project(db_session, "seven")
    report = _write_report(db_session, project, make_proposal_form_json())

    xlsx_bytes = render_proposal_form_xlsx(report, project)
    wb = _load(xlsx_bytes)

    assert wb.sheetnames == EXPECTED_SHEET_NAMES


def test_grand_total_matches_input(db_session: Session):
    project = _project(db_session, "gt")
    payload = make_proposal_form_json()
    by_wc = payload["base_bid"]["by_work_category"]
    unattributed = payload["base_bid"]["unattributed"]
    expected_total = sum(w["subtotal"] for w in by_wc)
    if unattributed is not None:
        expected_total += unattributed["subtotal"]

    report = _write_report(db_session, project, payload)
    wb = _load(render_proposal_form_xlsx(report, project))

    ws = wb["Base Bid"]
    grand_total_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value == "GRAND TOTAL":
            grand_total_row = cell.row
            break
    assert grand_total_row is not None, "Grand total row missing"
    grand_total_cell = ws.cell(row=grand_total_row, column=6).value
    assert grand_total_cell == pytest.approx(expected_total)


def test_unattributed_bucket_marked_with_warning(db_session: Session):
    project = _project(db_session, "unattr")
    payload = make_proposal_form_json(
        base_bid={
            "total": 1500.0,
            "by_work_category": [
                {
                    "wc_number": "10",
                    "wc_title": "Earthwork",
                    "line_items_count": 2,
                    "labor_cost": 500.0,
                    "material_cost": 500.0,
                    "subtotal": 1000.0,
                    "attribution_confidence_avg": 0.9,
                }
            ],
            "unattributed": {
                "line_items_count": 1,
                "labor_cost": 250.0,
                "material_cost": 250.0,
                "subtotal": 500.0,
                "note": "Takeoff items with no matching WorkCategory — review before submission",
            },
        },
        warnings=[f"{WARN_UNATTRIBUTED}: 1/3 takeoff items unattributed (33.3%)"],
    )
    report = _write_report(db_session, project, payload)

    wb = _load(render_proposal_form_xlsx(report, project))
    ws = wb["Base Bid"]

    unattr_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.startswith("⚠ UNATTRIBUTED"):
            unattr_row = cell.row
            break
    assert unattr_row is not None, "Unattributed row missing"
    assert ws.cell(row=unattr_row, column=1).fill.fgColor.rgb == WARN_FILL.fgColor.rgb


def test_alternates_no_price_highlighted(db_session: Session):
    project = _project(db_session, "alts")
    payload = make_proposal_form_json(
        alternates=[
            {
                "wc_number": "10",
                "description": "Missing price",
                "price_type": "add",
                "amount": None,
                "source": "work_category.add_alternates",
            },
            {
                "wc_number": "10",
                "description": "Priced",
                "price_type": "add",
                "amount": 1000.0,
                "source": "work_category.add_alternates",
            },
        ],
        warnings=[f"{WARN_ALTS_NO_PRICE}: 1 alternate has no price set"],
    )
    report = _write_report(db_session, project, payload)

    wb = _load(render_proposal_form_xlsx(report, project))
    ws = wb["Alternates"]

    # Row 2 has no amount → highlighted. Row 3 has amount → not highlighted.
    row2_fill = ws.cell(row=2, column=1).fill.fgColor.rgb
    row3_fill = ws.cell(row=3, column=1).fill.fgColor.rgb
    assert row2_fill == WARN_FILL.fgColor.rgb
    assert row3_fill != WARN_FILL.fgColor.rgb


def test_empty_alternates_shows_none_row(db_session: Session):
    project = _project(db_session, "empty")
    payload = make_proposal_form_json(alternates=[])
    report = _write_report(db_session, project, payload)

    wb = _load(render_proposal_form_xlsx(report, project))
    ws = wb["Alternates"]
    assert ws.cell(row=2, column=1).value == "(none)"


# ---------------------------------------------------------------------------
# Endpoint tests
# ---------------------------------------------------------------------------
def test_endpoint_404_when_no_report(client, db_session, test_user, auth_headers):
    project = Project(
        name="No Report Project",
        project_number=f"NR-{uuid.uuid4().hex[:8]}",
        project_type="commercial",
        owner_id=test_user.id,
    )
    db_session.add(project)
    db_session.commit()
    db_session.refresh(project)

    resp = client.get(
        f"/api/projects/{project.id}/proposal-form/export.xlsx",
        headers=auth_headers,
    )
    assert resp.status_code == 404
    assert "intelligence report" in resp.json()["detail"].lower()


def test_endpoint_404_when_proposal_form_json_null(
    client, db_session, test_user, auth_headers
):
    project = Project(
        name="Null PF Project",
        project_number=f"NP-{uuid.uuid4().hex[:8]}",
        project_type="commercial",
        owner_id=test_user.id,
    )
    db_session.add(project)
    db_session.commit()
    db_session.refresh(project)
    _write_report(db_session, project, None)

    resp = client.get(
        f"/api/projects/{project.id}/proposal-form/export.xlsx",
        headers=auth_headers,
    )
    assert resp.status_code == 404
    assert "not yet generated" in resp.json()["detail"].lower()


def test_endpoint_returns_xlsx_content_type_and_disposition(
    client, db_session, test_user, auth_headers
):
    project = Project(
        name="Happy Path",
        project_number=f"HP-{uuid.uuid4().hex[:8]}",
        project_type="commercial",
        owner_id=test_user.id,
    )
    db_session.add(project)
    db_session.commit()
    db_session.refresh(project)
    _write_report(db_session, project, make_proposal_form_json())

    resp = client.get(
        f"/api/projects/{project.id}/proposal-form/export.xlsx",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disp = resp.headers["content-disposition"]
    assert disp.startswith("attachment; filename=")
    assert "Proposal_Form_" in disp
    assert disp.endswith('.xlsx"')

    # Sanity: bytes are a real workbook with 7 sheets.
    wb = _load(resp.content)
    assert wb.sheetnames == EXPECTED_SHEET_NAMES


def test_filename_sanitization():
    fname = build_export_filename("AWS DCDE / Phase 2: KCCU")
    assert "/" not in fname
    assert ":" not in fname
    assert '"' not in fname
    # No whitespace runs
    assert "  " not in fname
    assert " " not in fname
    # Has the standard suffix shape
    assert fname.endswith(".xlsx")
    assert "_Proposal_Form_" in fname
