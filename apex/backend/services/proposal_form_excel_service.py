"""Sprint 18.4.3 — Proposal Form Excel export service.

Renders ``proposal_form_json`` from an IntelligenceReport into a 7-sheet
GC-ready workbook. Pure deterministic Python — no LLM imports, no IO
beyond bytes return. Totals/counts are computed in Python rather than
xlsx formulas so output is stable across Excel versions.

The 18.4.2 payload carries WorkCategory-level rollups (not raw line
items); the Base Bid sheet reflects that shape. PARKED: when Agent 6
densifies to per-item rows, the Base Bid sheet is the place to expand.
"""

from __future__ import annotations

import json
import logging
import os
import re
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apex.backend.models.intelligence_report import IntelligenceReportModel
from apex.backend.models.project import Project

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="FFD9D9D9")
WARN_FILL = PatternFill("solid", fgColor="FFFFE699")
BOLD_FONT = Font(bold=True)
CURRENCY_FMT = '"$"#,##0.00'
_TOP_BORDER = Border(top=Side(style="thin"))
_MAX_COL_WIDTH = 50

WARN_UNATTRIBUTED = "UNATTRIBUTED_ITEMS"
WARN_ALTS_NO_PRICE = "ALTERNATES_NO_PRICE"
WARN_UP_PLACEHOLDER = "UNIT_PRICES_PLACEHOLDER"
WARN_ALLOW_NO_AMT = "ALLOWANCES_NO_AMOUNT"
WARN_WC_EMPTY = "WC_EMPTY"

_FILENAME_UNSAFE = re.compile(r'[/\\:*?"<>|]')
_WHITESPACE_RUN = re.compile(r"\s+")


def render_proposal_form_xlsx(
    report: IntelligenceReportModel,
    project: Project,
) -> bytes:
    """Return xlsx bytes. Raises ValueError when proposal_form_json is empty."""
    raw = report.proposal_form_json
    if not raw:
        raise ValueError("proposal_form_json is empty; cannot render workbook")
    try:
        payload: dict[str, Any] = json.loads(raw) if isinstance(raw, str) else dict(raw)
    except (TypeError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"proposal_form_json is not valid JSON: {exc}") from exc

    wb = Workbook()
    wb.remove(wb.active)

    warnings = list(payload.get("warnings") or [])
    prefixes = {w.split(":", 1)[0] for w in warnings if ":" in w}

    _render_summary_sheet(wb, project, payload, warnings)
    _render_base_bid_sheet(wb, payload, prefixes)
    _render_alternates_sheet(wb, payload, prefixes)
    _render_allowances_sheet(wb, payload, prefixes)
    _render_unit_prices_sheet(wb, payload, prefixes)
    _render_breakouts_sheet(wb, payload)
    _render_warnings_sheet(wb, warnings)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export_filename(project_name: str, now: datetime | None = None) -> str:
    """Sanitize project name → ``<safe>_Proposal_Form_YYYYMMDD.xlsx``."""
    stamp = (now or datetime.now(UTC)).strftime("%Y%m%d")
    safe = _FILENAME_UNSAFE.sub("_", project_name or "project")
    safe = _WHITESPACE_RUN.sub("_", safe).strip("_") or "project"
    return f"{safe[:80]}_Proposal_Form_{stamp}.xlsx"


# ---------------------------------------------------------------------------
# Sheet 1 — Summary
# ---------------------------------------------------------------------------
def _render_summary_sheet(
    wb: Workbook, project: Project, payload: dict[str, Any], warnings: list[str]
) -> None:
    ws = wb.create_sheet("Summary")
    base_bid = payload.get("base_bid") or {}
    total = float(base_bid.get("total") or 0.0)

    meta = [
        ("Project Name", project.name),
        ("Project Number", project.project_number),
        ("Bid Date", project.bid_date or project.bid_due_date or ""),
        ("General Contractor", getattr(project, "gc_name", "") or ""),
        ("", ""),
        ("Generated At", datetime.now(UTC).isoformat(timespec="seconds")),
        ("APEX Version", os.environ.get("APEX_VERSION", "0.12.0")),
    ]
    for i, (label, value) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws.cell(row=i, column=2, value=value)

    start = len(meta) + 2
    ws.cell(row=start, column=1, value="Totals").font = BOLD_FONT
    totals = [
        ("Base Bid Total", total, CURRENCY_FMT),
        ("Alternates (count)", len(payload.get("alternates") or []), None),
        ("Allowances (count)", len(payload.get("allowances") or []), None),
        ("Unit Prices (count)", len(payload.get("unit_prices") or []), None),
        ("Warnings (count)", len(warnings), None),
    ]
    for i, (label, value, fmt) in enumerate(totals, start=start + 1):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=value)
        if fmt:
            c.number_format = fmt

    final = start + len(totals) + 2
    ws.cell(row=final, column=1, value="Total Bid Value").font = BOLD_FONT
    t = ws.cell(row=final, column=2, value=total)
    t.font = BOLD_FONT
    t.number_format = CURRENCY_FMT

    _autosize(ws, (1, 2))
    _apply_print_setup(ws)


# ---------------------------------------------------------------------------
# Sheet 2 — Base Bid
# ---------------------------------------------------------------------------
def _render_base_bid_sheet(
    wb: Workbook, payload: dict[str, Any], prefixes: set[str]
) -> None:
    ws = wb.create_sheet("Base Bid")
    headers = ["Work Category", "Title", "Line Items", "Labor Cost", "Material Cost", "Subtotal"]
    _write_header(ws, headers)

    base_bid = payload.get("base_bid") or {}
    by_wc = base_bid.get("by_work_category") or []
    unattributed = base_bid.get("unattributed")

    row = 2
    grand_total = 0.0
    for wc in by_wc:
        ws.cell(row=row, column=1, value=wc.get("wc_number", ""))
        ws.cell(row=row, column=2, value=wc.get("wc_title", ""))
        ws.cell(row=row, column=3, value=int(wc.get("line_items_count") or 0))
        _money_cell(ws, row, 4, wc.get("labor_cost"))
        _money_cell(ws, row, 5, wc.get("material_cost"))
        _money_cell(ws, row, 6, wc.get("subtotal")).font = BOLD_FONT
        grand_total += float(wc.get("subtotal") or 0.0)
        row += 1

    if unattributed:
        label = ws.cell(
            row=row, column=1,
            value=f"⚠ UNATTRIBUTED — {unattributed.get('note') or 'Unattributed Items'}",
        )
        label.font = BOLD_FONT
        ws.cell(row=row, column=3, value=int(unattributed.get("line_items_count") or 0))
        _money_cell(ws, row, 4, unattributed.get("labor_cost"))
        _money_cell(ws, row, 5, unattributed.get("material_cost"))
        _money_cell(ws, row, 6, unattributed.get("subtotal")).font = BOLD_FONT
        grand_total += float(unattributed.get("subtotal") or 0.0)
        if WARN_UNATTRIBUTED in prefixes:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = WARN_FILL
        row += 1

    gt_label = ws.cell(row=row, column=1, value="GRAND TOTAL")
    gt_label.font = BOLD_FONT
    gt_label.border = _TOP_BORDER
    for col in range(2, len(headers)):
        ws.cell(row=row, column=col).border = _TOP_BORDER
    gt = _money_cell(ws, row, 6, round(grand_total, 2))
    gt.font = BOLD_FONT
    gt.border = _TOP_BORDER

    _autosize(ws, range(1, len(headers) + 1))
    _apply_print_setup(ws)


# ---------------------------------------------------------------------------
# Sheets 3–5 — Alternates / Allowances / Unit Prices
# ---------------------------------------------------------------------------
def _render_alternates_sheet(
    wb: Workbook, payload: dict[str, Any], prefixes: set[str]
) -> None:
    ws = wb.create_sheet("Alternates")
    headers = ["Alt #", "Description", "Add/Deduct", "Amount", "Notes"]
    _write_header(ws, headers)
    alts = payload.get("alternates") or []
    if not alts:
        _none_row(ws, headers)
        return
    warn_on = WARN_ALTS_NO_PRICE in prefixes
    for idx, alt in enumerate(alts, start=1):
        r = idx + 1
        amt = alt.get("amount")
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=alt.get("description", ""))
        ws.cell(row=r, column=3, value=alt.get("price_type", ""))
        _money_cell(ws, r, 4, amt)
        ws.cell(row=r, column=5, value=alt.get("source", ""))
        if warn_on and amt is None:
            _fill_row(ws, r, len(headers), WARN_FILL)
    _autosize(ws, range(1, len(headers) + 1))
    _apply_print_setup(ws)


def _render_allowances_sheet(
    wb: Workbook, payload: dict[str, Any], prefixes: set[str]
) -> None:
    ws = wb.create_sheet("Allowances")
    headers = ["Allowance #", "Description", "Amount", "Notes"]
    _write_header(ws, headers)
    allowances = payload.get("allowances") or []
    if not allowances:
        _none_row(ws, headers)
        return
    warn_on = WARN_ALLOW_NO_AMT in prefixes
    for idx, allow in enumerate(allowances, start=1):
        r = idx + 1
        amt = allow.get("amount")
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=allow.get("description", ""))
        _money_cell(ws, r, 3, amt)
        ws.cell(row=r, column=4, value=allow.get("source", ""))
        if warn_on and amt is None:
            _fill_row(ws, r, len(headers), WARN_FILL)
    _autosize(ws, range(1, len(headers) + 1))
    _apply_print_setup(ws)


def _render_unit_prices_sheet(
    wb: Workbook, payload: dict[str, Any], prefixes: set[str]
) -> None:
    ws = wb.create_sheet("Unit Prices")
    headers = ["Item", "UOM", "Unit Price", "Notes"]
    _write_header(ws, headers)
    prices = payload.get("unit_prices") or []
    if not prices:
        _none_row(ws, headers)
        return
    warn_on = WARN_UP_PLACEHOLDER in prefixes
    for idx, up in enumerate(prices, start=1):
        r = idx + 1
        rate = up.get("rate")
        ws.cell(row=r, column=1, value=up.get("description", ""))
        ws.cell(row=r, column=2, value=up.get("unit", ""))
        _money_cell(ws, r, 3, rate)
        ws.cell(row=r, column=4, value=up.get("source", ""))
        if warn_on and rate is None:
            _fill_row(ws, r, len(headers), WARN_FILL)
    _autosize(ws, range(1, len(headers) + 1))
    _apply_print_setup(ws)


# ---------------------------------------------------------------------------
# Sheet 6 — Inclusions & Exclusions
# ---------------------------------------------------------------------------
def _render_breakouts_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("Inclusions & Exclusions")
    notes = payload.get("breakout_notes") or []
    groups: dict[str, list[dict[str, Any]]] = {"Inclusions": [], "Exclusions": [], "Clarifications": []}
    for note in notes:
        desc = (note.get("description") or "").lower()
        if "exclu" in desc or "not-to-exceed" in desc or "not to exceed" in desc:
            groups["Exclusions"].append(note)
        elif "clarif" in desc:
            groups["Clarifications"].append(note)
        else:
            groups["Inclusions"].append(note)

    row = 1
    for section in ("Inclusions", "Exclusions", "Clarifications"):
        h = ws.cell(row=row, column=1, value=section)
        h.font = BOLD_FONT
        h.fill = HEADER_FILL
        ws.cell(row=row, column=2, value="WC").font = BOLD_FONT
        ws.cell(row=row, column=3, value="Description").font = BOLD_FONT
        row += 1
        items = groups[section]
        if not items:
            ws.cell(row=row, column=1, value="(none)")
            row += 2
            continue
        for n in sorted(items, key=lambda x: x.get("wc_number") or ""):
            ws.cell(row=row, column=2, value=n.get("wc_number", ""))
            ws.cell(row=row, column=3, value=n.get("description", ""))
            row += 1
        row += 1

    ws.freeze_panes = "A2"
    _autosize(ws, (1, 2, 3))
    _apply_print_setup(ws)


# ---------------------------------------------------------------------------
# Sheet 7 — Warnings & QC
# ---------------------------------------------------------------------------
def _render_warnings_sheet(wb: Workbook, warnings: list[str]) -> None:
    ws = wb.create_sheet("Warnings & QC")
    if not warnings:
        ws.cell(row=1, column=1, value="0 warnings").font = BOLD_FONT
        _write_header(ws, ["Warning Type", "Description", "Affected Section"], row=2)
        ws.cell(row=3, column=1, value="(none)")
        _autosize(ws, (1, 2, 3))
        _apply_print_setup(ws)
        return

    counts: dict[str, int] = {}
    parsed: list[tuple[str, str]] = []
    for w in warnings:
        prefix, _, rest = w.partition(":")
        prefix = prefix.strip()
        parsed.append((prefix, rest.strip() or w))
        counts[prefix] = counts.get(prefix, 0) + 1

    summary = ", ".join(f"{k}×{v}" for k, v in sorted(counts.items()))
    top = ws.cell(row=1, column=1, value=f"{len(warnings)} warnings: {summary}")
    top.font = BOLD_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=3)
    _write_header(ws, ["Warning Type", "Description", "Affected Section"], row=2)
    for idx, (prefix, detail) in enumerate(parsed, start=3):
        ws.cell(row=idx, column=1, value=prefix).fill = WARN_FILL
        ws.cell(row=idx, column=2, value=detail)
        ws.cell(row=idx, column=3, value=_affected_section(prefix))
    _autosize(ws, (1, 2, 3))
    _apply_print_setup(ws)


def _affected_section(prefix: str) -> str:
    return {
        WARN_UNATTRIBUTED: "Base Bid",
        WARN_ALTS_NO_PRICE: "Alternates",
        WARN_UP_PLACEHOLDER: "Unit Prices",
        WARN_ALLOW_NO_AMT: "Allowances",
        WARN_WC_EMPTY: "Base Bid",
    }.get(prefix, "")


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _write_header(ws: Worksheet, headers: list[str], *, row: int = 1) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = BOLD_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _none_row(ws: Worksheet, headers: list[str]) -> None:
    ws.cell(row=2, column=1, value="(none)")
    _autosize(ws, range(1, len(headers) + 1))
    _apply_print_setup(ws)


def _money_cell(ws: Worksheet, row: int, col: int, value: Any):
    cell = ws.cell(row=row, column=col, value=value if value is not None else "")
    cell.number_format = CURRENCY_FMT
    return cell


def _fill_row(ws: Worksheet, row: int, n_cols: int, fill: PatternFill) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def _autosize(ws: Worksheet, cols) -> None:
    for col_idx in cols:
        max_len = 10
        letter = get_column_letter(col_idx)
        for cell in ws[letter]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + 2, _MAX_COL_WIDTH)


def _apply_print_setup(ws: Worksheet) -> None:
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
