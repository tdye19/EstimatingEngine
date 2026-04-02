"""Parse estimator WinEst takeoff uploads into structured TakeoffLineItem lists.

Reuses format detection patterns from productivity_brain.parser but produces
TakeoffLineItem objects (pipeline_contracts) instead of PBLineItem DB records.
The key difference: PB parser ingests historical data for averaging.
Takeoff parser ingests THIS bid's quantities for rate comparison.

Supported formats:
- WinEst .est (OLE2 native) -- binary; extracts descriptions, numeric fields may be None
- WinEst 26-col (CCI Civil Est Report) -- cell A1 contains "_CCI Civil Est Report"
- WinEst 21-col (CCI Estimate Report) -- cell A1 contains "_CCI Estimate Report"
- Simple CSV/XLSX -- header row with columns: Activity, Qty, Unit, Crew, Rate
- Generic takeoff -- any spreadsheet with an activity/description column + quantity columns
"""

import csv
import os
from typing import Optional

import openpyxl
import pandas as pd

from apex.backend.agents.pipeline_contracts import TakeoffLineItem


# ── Helpers ──────────────────────────────────────────────────────────────────

def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return None
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _safe_str(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    s = str(val).strip()
    return s if s and s.lower() != "nan" else None


# ── Format detection ─────────────────────────────────────────────────────────

def detect_takeoff_format(filepath: str) -> str:
    """Detect takeoff file format.

    Returns "26col", "21col", "simple_csv", or "unknown".
    """
    ext = os.path.splitext(filepath)[1].lower()

    # CSV files — check for simple CSV format
    if ext == ".csv":
        return _detect_csv_format(filepath)

    # Excel files — check A1 cell for WinEst signatures
    try:
        df = pd.read_excel(filepath, header=None, nrows=5)
    except Exception:
        return "unknown"

    cell_a1 = str(df.iloc[0, 0]) if pd.notna(df.iloc[0, 0]) else ""

    if "_CCI Civil Est Report" in cell_a1 or "CCI Civil Est Report" in cell_a1:
        return "26col"
    if "_CCI Estimate Report" in cell_a1 or "CCI Estimate Report" in cell_a1:
        return "21col"

    # Not WinEst — check if it's a simple xlsx with Activity + Qty headers
    return _detect_simple_header(df)


def _detect_csv_format(filepath: str) -> str:
    """Check if a CSV has Activity + Qty/Quantity headers."""
    try:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 5:
                    break
                lower_row = [c.strip().lower() for c in row]
                if "activity" in lower_row and any(
                    h in lower_row for h in ("qty", "quantity")
                ):
                    return "simple_csv"
    except Exception:
        pass
    return "unknown"


def _detect_simple_header(df: pd.DataFrame) -> str:
    """Scan first 5 rows of a DataFrame for Activity + Qty/Quantity columns."""
    for i in range(min(5, len(df))):
        row_vals = [str(c).strip().lower() for c in df.iloc[i] if pd.notna(c)]
        if "activity" in row_vals and any(
            h in row_vals for h in ("qty", "quantity")
        ):
            return "simple_csv"
    return "unknown"


# ── 26-column parser ─────────────────────────────────────────────────────────

def parse_26col_takeoff(filepath: str) -> list[TakeoffLineItem]:
    """Parse a 26-column CCI Civil Est Report export.

    Header at row 5 (index 4). Column mapping:
      B -> wbs_area, C -> activity, F -> quantity, G -> unit,
      H -> crew, I -> production_rate (Unit/MH),
      K -> labor_cost_per_unit, L -> material_cost_per_unit
    """
    df = pd.read_excel(filepath, header=None)
    items: list[TakeoffLineItem] = []
    row_num = 0

    for idx in range(5, len(df)):  # row 5 is header (index 4), data starts at index 5
        row = df.iloc[idx]
        activity = _safe_str(row[2])  # col C (0-indexed)

        # Skip empty rows and section headers (start with digit)
        if activity is None:
            continue
        if activity[0].isdigit():
            continue

        row_num += 1
        items.append(TakeoffLineItem(
            row_number=row_num,
            wbs_area=_safe_str(row[1]),               # col B
            activity=activity,                          # col C
            quantity=_safe_float(row[5]),               # col F
            unit=_safe_str(row[6]),                     # col G
            crew=_safe_str(row[7]),                     # col H
            production_rate=_safe_float(row[8]),        # col I
            labor_cost_per_unit=_safe_float(row[10]),   # col K
            material_cost_per_unit=_safe_float(row[11]),  # col L
        ))

    return items


# ── 21-column parser ─────────────────────────────────────────────────────────

def parse_21col_takeoff(filepath: str) -> list[TakeoffLineItem]:
    """Parse a 21-column CCI Estimate Report export.

    Header at row 5 (index 4). Column mapping:
      B -> wbs_area, C -> activity, D -> quantity, E -> unit,
      F -> crew, G -> production_rate
    """
    df = pd.read_excel(filepath, header=None)
    items: list[TakeoffLineItem] = []
    row_num = 0

    for idx in range(5, len(df)):  # row 5 is header (index 4), data starts at index 5
        row = df.iloc[idx]
        activity = _safe_str(row[2])  # col C (0-indexed)

        if activity is None:
            continue
        if activity[0].isdigit():
            continue

        row_num += 1
        items.append(TakeoffLineItem(
            row_number=row_num,
            wbs_area=_safe_str(row[1]),           # col B
            activity=activity,                     # col C
            quantity=_safe_float(row[3]),          # col D
            unit=_safe_str(row[4]),                # col E
            crew=_safe_str(row[5]),                # col F
            production_rate=_safe_float(row[6]),   # col G
        ))

    return items


# ── Simple CSV/XLSX parser ────────────────────────────────────────────────────

def parse_simple_csv(filepath: str) -> list[TakeoffLineItem]:
    """Parse a simple CSV or XLSX with named columns.

    Scans first 5 rows for a header containing "Activity". Maps columns
    case-insensitively:
      Activity -> activity
      Qty / Quantity -> quantity
      Unit -> unit
      Crew -> crew
      Rate / Prod Rate / Production Rate -> production_rate
      WBS / WBS Area -> wbs_area
      CSI / CSI Code -> csi_code
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        return _parse_csv_file(filepath)
    else:
        return _parse_xlsx_simple(filepath)


def _find_header_row(rows: list[list]) -> tuple[Optional[int], dict[str, int]]:
    """Find the header row and build a column-name-to-index map."""
    col_aliases = {
        "activity": "activity",
        "qty": "quantity",
        "quantity": "quantity",
        "unit": "unit",
        "crew": "crew",
        "rate": "production_rate",
        "prod rate": "production_rate",
        "production rate": "production_rate",
        "wbs": "wbs_area",
        "wbs area": "wbs_area",
        "csi": "csi_code",
        "csi code": "csi_code",
    }

    for row_idx, row in enumerate(rows):
        lower_cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "activity" not in lower_cells:
            continue

        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(lower_cells):
            field = col_aliases.get(cell)
            if field and field not in col_map:
                col_map[field] = col_idx

        if "activity" in col_map:
            return row_idx, col_map

    return None, {}


def _parse_csv_file(filepath: str) -> list[TakeoffLineItem]:
    """Parse a CSV file with named columns."""
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    # Find header in first 5 rows
    header_idx, col_map = _find_header_row(all_rows[:5])
    if header_idx is None:
        return []

    items: list[TakeoffLineItem] = []
    row_num = 0

    for row in all_rows[header_idx + 1:]:
        activity_idx = col_map["activity"]
        if activity_idx >= len(row):
            continue
        activity = row[activity_idx].strip() if row[activity_idx] else None
        if not activity:
            continue

        row_num += 1
        items.append(TakeoffLineItem(
            row_number=row_num,
            activity=activity,
            quantity=_safe_float(row[col_map["quantity"]]) if "quantity" in col_map and col_map["quantity"] < len(row) else None,
            unit=row[col_map["unit"]].strip() if "unit" in col_map and col_map["unit"] < len(row) and row[col_map["unit"]].strip() else None,
            crew=row[col_map["crew"]].strip() if "crew" in col_map and col_map["crew"] < len(row) and row[col_map["crew"]].strip() else None,
            production_rate=_safe_float(row[col_map["production_rate"]]) if "production_rate" in col_map and col_map["production_rate"] < len(row) else None,
            wbs_area=row[col_map["wbs_area"]].strip() if "wbs_area" in col_map and col_map["wbs_area"] < len(row) and row[col_map["wbs_area"]].strip() else None,
            csi_code=row[col_map["csi_code"]].strip() if "csi_code" in col_map and col_map["csi_code"] < len(row) and row[col_map["csi_code"]].strip() else None,
        ))

    return items


def _parse_xlsx_simple(filepath: str) -> list[TakeoffLineItem]:
    """Parse an XLSX file with named columns using openpyxl."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Read first 5 rows to find header
    preview_rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 5:
            break
        preview_rows.append(list(row))

    header_idx, col_map = _find_header_row(preview_rows)
    if header_idx is None:
        wb.close()
        return []

    items: list[TakeoffLineItem] = []
    row_num = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue

        row_list = list(row)
        activity_idx = col_map["activity"]
        if activity_idx >= len(row_list):
            continue
        activity = _safe_str(row_list[activity_idx])
        if not activity:
            continue

        row_num += 1

        def _get(field: str) -> Optional:
            idx = col_map.get(field)
            if idx is None or idx >= len(row_list):
                return None
            return row_list[idx]

        items.append(TakeoffLineItem(
            row_number=row_num,
            activity=activity,
            quantity=_safe_float(_get("quantity")),
            unit=_safe_str(_get("unit")),
            crew=_safe_str(_get("crew")),
            production_rate=_safe_float(_get("production_rate")),
            wbs_area=_safe_str(_get("wbs_area")),
            csi_code=_safe_str(_get("csi_code")),
        ))

    wb.close()
    return items


# ── Generic takeoff parser ──────────────────────────────────────────────────

# Column name matching (case-insensitive, stripped)
_ACTIVITY_KEYWORDS = {
    "label", "description", "activity", "item", "scope", "work item",
    "work description", "element", "component", "task",
}

_QUANTITY_KEYWORDS = {
    # Direct unit names
    "qty", "quantity", "amount",
    # Linear
    "lf", "lnft", "lineal ft", "linear feet", "total lnft", "length",
    # Area
    "sf", "sqft", "area", "total area", "total area (sf)", "square feet",
    # Volume
    "cy", "cuyd", "cubic yards", "total cuyd", "volume", "total cy",
    # Count
    "ea", "each", "count", "pcs", "pieces",
    # Weight
    "tons", "lbs", "pounds", "weight",
}

_UNIT_INFERENCE: dict[str, Optional[str]] = {
    "lf": "LF", "lnft": "LF", "lineal ft": "LF", "linear feet": "LF",
    "total lnft": "LF", "length": "LF",
    "sf": "SF", "sqft": "SF", "area": "SF", "total area": "SF",
    "total area (sf)": "SF", "square feet": "SF",
    "cy": "CY", "cuyd": "CY", "cubic yards": "CY", "total cuyd": "CY",
    "volume": "CY", "total cy": "CY",
    "ea": "EA", "each": "EA", "count": "EA", "pcs": "EA", "pieces": "EA",
    "tons": "TON", "lbs": "LBS", "pounds": "LBS", "weight": "LBS",
    "qty": None, "quantity": None, "amount": None,
}

# Priority order for picking the primary quantity column
_UNIT_PRIORITY = ["CY", "SF", "LF", "EA", "TON", "LBS", None]

_TOTAL_KEYWORDS = {"total", "subtotal", "sum", "grand total"}


def _match_keyword(cell_text: str, keywords: set[str]) -> Optional[str]:
    """Check if cell_text matches any keyword (exact or substring).

    Returns the matched keyword, or None.
    """
    lower = cell_text.strip().lower()
    if not lower:
        return None
    # Exact match first
    if lower in keywords:
        return lower
    # Substring match — keyword appears in cell value
    for kw in keywords:
        if kw in lower:
            return kw
    return None


def _detect_generic_header(rows: list[list]) -> tuple[Optional[int], Optional[int], list[tuple[int, str, Optional[str]]]]:
    """Scan rows for a generic takeoff header.

    Returns (header_row_index, activity_col_index, qty_columns) where
    qty_columns is a list of (col_index, matched_keyword, inferred_unit).
    Returns (None, None, []) if no suitable header found.
    """
    for row_idx, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]

        activity_col = None
        qty_cols: list[tuple[int, str, Optional[str]]] = []

        for col_idx, cell in enumerate(cells):
            if not cell:
                continue
            # Check activity
            act_match = _match_keyword(cell, _ACTIVITY_KEYWORDS)
            if act_match is not None and activity_col is None:
                activity_col = col_idx
                continue
            # Check quantity
            qty_match = _match_keyword(cell, _QUANTITY_KEYWORDS)
            if qty_match is not None:
                unit = _UNIT_INFERENCE.get(qty_match)
                qty_cols.append((col_idx, qty_match, unit))

        if activity_col is not None and len(qty_cols) >= 1:
            return row_idx, activity_col, qty_cols

    return None, None, []


def _pick_primary_qty(qty_cols: list[tuple[int, str, Optional[str]]], row_data: list) -> tuple[Optional[float], Optional[str]]:
    """Pick the best quantity column value for a data row.

    Priority: CY > SF > LF > EA > TON > LBS > generic (None unit).
    Within a priority level, picks the first column with a non-None value.
    """
    # Group columns by inferred unit
    by_unit: dict[Optional[str], list[tuple[int, str]]] = {}
    for col_idx, kw, unit in qty_cols:
        by_unit.setdefault(unit, []).append((col_idx, kw))

    for target_unit in _UNIT_PRIORITY:
        cols = by_unit.get(target_unit, [])
        for col_idx, _kw in cols:
            if col_idx < len(row_data):
                val = _safe_float(row_data[col_idx])
                if val is not None:
                    return val, target_unit
    return None, None


def _is_total_row(activity: str) -> bool:
    """Return True if activity looks like a total/summary row."""
    lower = activity.strip().lower()
    return any(lower == kw or lower.startswith(kw + " ") or lower.startswith(kw + ":")
               for kw in _TOTAL_KEYWORDS)


def _try_generic_format_ws(ws) -> Optional[dict]:
    """Try to parse a single worksheet as a generic takeoff.

    Returns {"items": [...], "warnings": [...]} or None if no header found.
    """
    # Read first 10 rows for header detection
    preview: list[list] = []
    all_rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_list = list(row)
        all_rows.append(row_list)
        if i < 10:
            preview.append(row_list)

    header_idx, activity_col, qty_cols = _detect_generic_header(preview)
    if header_idx is None:
        return None

    items: list[TakeoffLineItem] = []
    warnings: list[str] = []
    row_num = 0

    for i, row_data in enumerate(all_rows):
        if i <= header_idx:
            continue

        # Get activity
        if activity_col >= len(row_data):
            continue
        activity = _safe_str(row_data[activity_col])
        if not activity:
            continue  # skip blank rows
        if _is_total_row(activity):
            continue  # skip total/summary rows

        # Pick primary quantity
        qty_val, unit = _pick_primary_qty(qty_cols, row_data)

        row_num += 1
        items.append(TakeoffLineItem(
            row_number=row_num,
            activity=activity,
            quantity=qty_val,
            unit=unit,
        ))

    if not items:
        return None

    return {"items": items, "warnings": warnings}


def _try_generic_format(filepath: str) -> Optional[dict]:
    """Try to parse an xlsx file as a generic takeoff.

    Tries the active sheet first, then other sheets.
    Returns {"items": [...], "warnings": [...]} or None.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None

    # Try active sheet first
    result = _try_generic_format_ws(wb.active)
    if result and result["items"]:
        wb.close()
        return result

    # Try other sheets
    for name in wb.sheetnames:
        ws = wb[name]
        if ws == wb.active:
            continue
        result = _try_generic_format_ws(ws)
        if result and result["items"]:
            result["warnings"].append(f"Parsed from sheet '{name}' (not the first sheet).")
            wb.close()
            return result

    wb.close()
    return None


# ── .est (OLE2 native) parser ───────────────────────────────────────────────

def _parse_est_file(filepath: str) -> list[TakeoffLineItem]:
    """Parse a native WinEst .est file using the existing OLE2 parser.

    Delegates to winest_parser.parse_winest_file() for binary extraction,
    then normalizes output into TakeoffLineItem objects.

    Items from .est files will typically have:
    - activity: extracted text description
    - All numeric fields: None (proprietary binary format)

    These items still get rate recommendations from Agent 4 via
    description-based fuzzy matching against Productivity Brain.
    """
    from apex.backend.utils.winest_parser import parse_winest_file

    result = parse_winest_file(filepath)

    if not result["success"]:
        return []

    items: list[TakeoffLineItem] = []
    for idx, raw_item in enumerate(result["line_items"], start=1):
        desc = (raw_item.get("description") or "").strip()
        if not desc or len(desc) < 3:
            continue  # skip noise

        items.append(TakeoffLineItem(
            row_number=idx,
            wbs_area=raw_item.get("wbs_code"),
            activity=desc,
            quantity=raw_item.get("quantity"),
            unit=raw_item.get("unit"),
            crew=_safe_str(raw_item.get("crew_size")),
            production_rate=raw_item.get("productivity_rate"),
            labor_cost_per_unit=raw_item.get("labor_rate"),
            material_cost_per_unit=raw_item.get("material_cost"),
            csi_code=raw_item.get("csi_code"),
        ))

    return items


# ── Main dispatcher ──────────────────────────────────────────────────────────

def parse_takeoff(filepath: str) -> tuple[list[TakeoffLineItem], str]:
    """Parse a takeoff file, auto-detecting format.

    Returns (items, format_name) where format_name is one of:
    "est_native", "26col", "21col", "simple_csv", "generic_takeoff", "unknown".
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".est":
        return _parse_est_file(filepath), "est_native"

    fmt = detect_takeoff_format(filepath)

    if fmt == "26col":
        return parse_26col_takeoff(filepath), fmt
    elif fmt == "21col":
        return parse_21col_takeoff(filepath), fmt
    elif fmt == "simple_csv":
        return parse_simple_csv(filepath), fmt
    else:
        # Try generic format for xlsx/xls before giving up
        if ext in (".xlsx", ".xls"):
            generic = _try_generic_format(filepath)
            if generic and generic["items"]:
                return generic["items"], "generic_takeoff"
        return [], fmt


# ── Quick smoke test ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python -m apex.backend.services.takeoff_parser.parser <filepath>")
        sys.exit(1)

    path = sys.argv[1]
    fmt = detect_takeoff_format(path)
    print(f"Detected format: {fmt}")

    items, detected = parse_takeoff(path)
    print(f"Parsed {len(items)} line items (format={detected})")

    for item in items[:5]:
        print(f"  #{item.row_number}: {item.activity} | qty={item.quantity} {item.unit} | rate={item.production_rate}")

    if len(items) > 5:
        print(f"  ... and {len(items) - 5} more")
