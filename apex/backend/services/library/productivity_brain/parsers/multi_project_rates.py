"""Parser for files that preserve per-project rate columns.

Input shape (row numbers are 1-indexed to match Excel):
  row 1   title banner
  row 2   subtitle banner
  row 3   blank
  row 4   header   → WBS Area | Activity Description | Unit | Crew / Trade |
                     AVG Prod | <1..N project cols> | Count | Spread |
                     Avg Labor $/Unit | Avg Mat $/Unit
  row 5+  alternating section-header rows (col A populated, rest empty) and
          activity rows (Activity Description populated)

One ParsedProject is emitted per per-project column. Each project's
line_items include only activities where that project's cell is a
non-sentinel numeric value. AVG Prod / Count / Spread are discarded —
PB recomputes its own aggregates at query time.
"""

from __future__ import annotations

import re
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel

# Sentinels are matched AFTER .strip().lower(). Whitespace-only also skipped.
_SENTINELS = {"", "—", "-", "--", "no hist data", "n/a", "tbd"}

# Reserved header-row tokens that are NOT project-name columns. Matched by
# substring on the lowercased header text.
_RESERVED_HEADER_SUBSTRINGS = ("avg prod", "count", "spread", "avg labor", "avg mat")

_SECTION_PREFIX_RE = re.compile(r"^\s*\d+\s*[—\-]+\s*")
_DEFAULT_PROJECT_NAME_TEMPLATE = "CCI CityGate {source_project}"
_MIN_EXPECTED_LINE_ITEMS = 40
# Only flag genuine unrelated drift between the current section header and
# the per-row col-0 sub-group. Sub-WBS labels (e.g. 'M & R Building - Walls'
# under '010 — M&R Building') share substantial tokens and score ~0.7+.
_WBS_DRIFT_RATIO_THRESHOLD = 0.3


class ParsedLineItem(BaseModel):
    wbs_area: str
    activity_description: str
    unit: str
    crew: str | None
    production_rate: float
    labor_cost_per_unit: float | None
    material_cost_per_unit: float | None
    csi_code: str | None = None


class ParsedProject(BaseModel):
    project_name: str
    source_project: str
    metadata: dict
    line_items: list[ParsedLineItem]


class ParseResult(BaseModel):
    parsed_projects: list[ParsedProject]
    warnings: list[str]


class _ColumnMap(BaseModel):
    wbs: int
    activity: int
    unit: int
    crew: int | None
    project_cols: list[tuple[str, int]]
    labor_up: int | None
    mat_up: int | None


def _coerce_cell(value: object) -> float | None:
    """Return float or None. Sentinels and unparseable strings map to None."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "" or stripped.lower() in _SENTINELS:
            return None
        try:
            return float(stripped.replace(",", "").replace("$", ""))
        except ValueError:
            return None
    return None


def _normalise_wbs(section_header: str) -> str:
    """Strip leading 'NNN — ' prefix for comparison to per-row col 0."""
    return _SECTION_PREFIX_RE.sub("", section_header).strip().lower()


class MultiProjectRatesParser:
    SHEET_NAME = "Averaged Prod Rates"

    @classmethod
    def detect(cls, file_path: str) -> bool:
        """True when the file has a `Averaged Prod Rates` sheet (or first sheet
        matching the layout) with a header row declaring WBS Area, Activity
        Description, Unit, AVG Prod, Count, and ≥2 project-name columns
        between AVG Prod and Count.
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            sheet = wb[cls.SHEET_NAME] if cls.SHEET_NAME in wb.sheetnames else wb.worksheets[0]
            header_row = cls._find_header_row(sheet)
            if header_row is None:
                return False
            col_map = cls._build_column_map(header_row)
            return col_map is not None and len(col_map.project_cols) >= 2
        finally:
            wb.close()

    def parse(
        self,
        file_path: str,
        metadata_overrides: dict | None = None,
        project_name_template: str = _DEFAULT_PROJECT_NAME_TEMPLATE,
    ) -> ParseResult:
        """Parse the file into per-project ParsedProject records.

        Raises ValueError if detection fails (don't silently fall through to
        another format).
        """
        if not self.detect(file_path):
            raise ValueError(
                f"File format not recognised as multi-project rates: {file_path}"
            )

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = (
                wb[self.SHEET_NAME]
                if self.SHEET_NAME in wb.sheetnames
                else wb.worksheets[0]
            )
            all_rows = list(sheet.iter_rows(values_only=True))
        finally:
            wb.close()

        header_row_idx = self._find_header_row_index(all_rows)
        if header_row_idx is None:
            raise ValueError("Header row not found after detection passed — unexpected")
        col_map = self._build_column_map(all_rows[header_row_idx])
        if col_map is None:
            raise ValueError("Column map could not be built after detection passed")

        base_metadata = {
            "file_source": Path(file_path).name,
            **(metadata_overrides or {}),
        }

        warnings: list[str] = []
        # source_project → list[ParsedLineItem]
        by_project: dict[str, list[ParsedLineItem]] = {
            name: [] for name, _ in col_map.project_cols
        }

        current_section: str | None = None

        for row_idx, row in enumerate(all_rows):
            if row_idx <= header_row_idx:
                continue
            # Normalise to tuple of length at least max needed
            col0 = row[col_map.wbs] if len(row) > col_map.wbs else None
            col1 = row[col_map.activity] if len(row) > col_map.activity else None

            # Section header: col 0 populated, cols 1..end all None
            if col0 is not None and all(v is None for v in row[1:]):
                current_section = str(col0).strip()
                continue

            # Activity row must have a description
            if col1 is None or not str(col1).strip():
                continue
            activity = str(col1).strip()

            if current_section is None:
                warnings.append(
                    f"row {row_idx + 1}: activity '{activity}' has no preceding section header; skipped"
                )
                continue

            # Per-row col 0 drift warning (non-fatal). Col 0 is *designed*
            # to carry a finer sub-WBS; only warn when it's unrelated to the
            # current section header.
            if col0 is not None:
                row_wbs_norm = str(col0).strip().lower()
                section_norm = _normalise_wbs(current_section)
                if row_wbs_norm and section_norm:
                    ratio = SequenceMatcher(None, row_wbs_norm, section_norm).ratio()
                    if ratio < _WBS_DRIFT_RATIO_THRESHOLD:
                        warnings.append(
                            f"row {row_idx + 1}: unrelated wbs — section='{current_section}' "
                            f"row_col0='{col0}' (similarity {ratio:.2f}; section used)"
                        )

            unit_val = row[col_map.unit] if len(row) > col_map.unit else None
            unit = str(unit_val).strip() if unit_val is not None else ""
            if not unit:
                warnings.append(
                    f"row {row_idx + 1}: activity '{activity}' has no unit; skipped"
                )
                continue

            crew_raw = (
                row[col_map.crew]
                if col_map.crew is not None and len(row) > col_map.crew
                else None
            )
            crew = str(crew_raw).strip() if crew_raw is not None and str(crew_raw).strip() else None

            labor_up = (
                _coerce_cell(row[col_map.labor_up])
                if col_map.labor_up is not None and len(row) > col_map.labor_up
                else None
            )
            mat_up = (
                _coerce_cell(row[col_map.mat_up])
                if col_map.mat_up is not None and len(row) > col_map.mat_up
                else None
            )

            for project_name, col_idx in col_map.project_cols:
                cell = row[col_idx] if len(row) > col_idx else None
                rate = _coerce_cell(cell)
                if rate is None:
                    continue
                by_project[project_name].append(
                    ParsedLineItem(
                        wbs_area=current_section,
                        activity_description=activity,
                        unit=unit,
                        crew=crew,
                        production_rate=rate,
                        labor_cost_per_unit=labor_up,
                        material_cost_per_unit=mat_up,
                        csi_code=None,
                    )
                )

        warnings.append(
            "labor_cost_per_unit and material_cost_per_unit are per-activity "
            "averages pooled across all projects in the source file — applied "
            "uniformly to each project's line items, not per-project specifics"
        )

        parsed_projects: list[ParsedProject] = []
        for name, _ in col_map.project_cols:
            items = by_project[name]
            if len(items) < _MIN_EXPECTED_LINE_ITEMS:
                warnings.append(
                    f"project '{name}' yielded only {len(items)} line items "
                    f"(expected ≥{_MIN_EXPECTED_LINE_ITEMS}); verify source file"
                )
            parsed_projects.append(
                ParsedProject(
                    project_name=project_name_template.format(source_project=name),
                    source_project=name,
                    metadata=dict(base_metadata),
                    line_items=items,
                )
            )

        return ParseResult(parsed_projects=parsed_projects, warnings=warnings)

    @classmethod
    def _find_header_row(cls, sheet) -> tuple | None:
        rows = list(sheet.iter_rows(min_row=1, max_row=15, values_only=True))
        idx = cls._find_header_row_index(rows)
        return rows[idx] if idx is not None else None

    @staticmethod
    def _find_header_row_index(rows: list) -> int | None:
        for i, row in enumerate(rows[:15]):
            if not row:
                continue
            texts = {str(v).strip().lower() for v in row if v is not None}
            if {"wbs area", "activity description", "unit"}.issubset(texts):
                return i
        return None

    @classmethod
    def _build_column_map(cls, header_row) -> _ColumnMap | None:
        if header_row is None:
            return None
        wbs = activity = unit = crew = avg_prod = count = labor_up = mat_up = None
        for i, v in enumerate(header_row):
            if v is None:
                continue
            text = str(v).strip().lower()
            if text == "wbs area":
                wbs = i
            elif text == "activity description":
                activity = i
            elif text == "unit":
                unit = i
            elif "crew" in text or "trade" in text:
                crew = i
            elif "avg prod" in text:
                avg_prod = i
            elif text == "count":
                count = i
            elif "avg labor" in text:
                labor_up = i
            elif "avg mat" in text:
                mat_up = i

        if wbs is None or activity is None or unit is None or avg_prod is None or count is None:
            return None

        project_cols: list[tuple[str, int]] = []
        for i in range(avg_prod + 1, count):
            v = header_row[i] if i < len(header_row) else None
            if v is None:
                continue
            text = str(v).strip()
            if not text:
                continue
            if any(sub in text.lower() for sub in _RESERVED_HEADER_SUBSTRINGS):
                continue
            project_cols.append((text, i))

        if len(project_cols) < 2:
            return None

        return _ColumnMap(
            wbs=wbs,
            activity=activity,
            unit=unit,
            crew=crew,
            project_cols=project_cols,
            labor_up=labor_up,
            mat_up=mat_up,
        )
