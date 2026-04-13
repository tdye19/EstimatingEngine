"""Estimate export router — PDF and XLSX downloads."""

import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from apex.backend.db.database import get_db
from apex.backend.models.agent_run_log import AgentRunLog
from apex.backend.models.estimate import Estimate
from apex.backend.models.gap_report import GapReport
from apex.backend.models.organization import Organization
from apex.backend.models.user import User
from apex.backend.utils.auth import get_authorized_project, require_auth

router = APIRouter(prefix="/api/exports", tags=["exports"])


def _get_estimate_or_404(project_id: int, db: Session, user: User):
    project = get_authorized_project(project_id, user, db)

    estimate = (
        db.query(Estimate)
        .filter(
            Estimate.project_id == project_id,
            Estimate.is_deleted == False,  # noqa: E712
        )
        .order_by(Estimate.version.desc())
        .first()
    )
    if not estimate:
        raise HTTPException(status_code=404, detail="No estimate available for this project")

    org = db.query(Organization).filter(Organization.id == project.organization_id).first()
    return project, estimate, org


# ─────────────────────────── PDF ───────────────────────────


@router.get("/projects/{project_id}/estimate/pdf")
def export_estimate_pdf(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    project, estimate, org = _get_estimate_or_404(project_id, db, current_user)

    # Fetch gap report and agent run logs for this project
    gap_report = (
        db.query(GapReport)
        .filter(
            GapReport.project_id == project_id,
            GapReport.is_deleted == False,  # noqa: E712
        )
        .order_by(GapReport.id.desc())
        .first()
    )

    agent_logs = (
        db.query(AgentRunLog)
        .filter(
            AgentRunLog.project_id == project_id,
        )
        .order_by(AgentRunLog.agent_number)
        .all()
    )

    buf = io.BytesIO()

    # ── Brand colours ──
    TCA_NAVY = colors.HexColor("#1E3A5F")
    TCA_NAVY_DARK = colors.HexColor("#152D4A")
    LIGHT_GRAY = colors.HexColor("#f1f5f9")
    MID_GRAY = colors.HexColor("#94a3b8")
    GRID_COLOR = colors.HexColor("#e2e8f0")
    SUBTOTAL_BG = colors.HexColor("#dbeafe")

    # ── Page number callback ──
    def _add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(MID_GRAY)
        page_w, _ = letter
        canvas.drawCentredString(page_w / 2, 0.5 * inch, f"Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=1 * inch,
        rightMargin=1 * inch,
        topMargin=1 * inch,
        bottomMargin=1 * inch,
    )

    styles = getSampleStyleSheet()
    style_cover_title = ParagraphStyle(
        "cover_title",
        parent=styles["Normal"],
        fontSize=28,
        fontName="Helvetica-Bold",
        textColor=TCA_NAVY,
        alignment=1,
        spaceAfter=8,
    )
    style_cover_sub = ParagraphStyle(
        "cover_sub",
        parent=styles["Normal"],
        fontSize=14,
        fontName="Helvetica",
        textColor=colors.HexColor("#475569"),
        alignment=1,
        spaceAfter=4,
    )
    style_cover_meta = ParagraphStyle(
        "cover_meta", parent=styles["Normal"], fontSize=10, textColor=MID_GRAY, alignment=1
    )
    style_section = ParagraphStyle(
        "section",
        parent=styles["Normal"],
        fontSize=12,
        fontName="Helvetica-Bold",
        textColor=TCA_NAVY,
        spaceBefore=14,
        spaceAfter=6,
    )
    style_body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9, leading=13)
    style_small = ParagraphStyle(
        "small", parent=styles["Normal"], fontSize=8, leading=11, textColor=colors.HexColor("#475569")
    )
    style_footer_text = ParagraphStyle(
        "footer_text", parent=styles["Normal"], fontSize=8, textColor=MID_GRAY, alignment=1
    )

    def fmt_dollar(val):
        return f"${val:,.0f}" if val is not None else "$0"

    def fmt_pct(val):
        return f"{val:.1f}%" if val is not None else "0.0%"

    # Usable table width
    usable_w = letter[0] - 2 * inch

    story = []

    # ══════════════════════════════════════════════════════════
    # PAGE 1: COVER PAGE
    # ══════════════════════════════════════════════════════════
    story.append(Spacer(1, 1.8 * inch))

    # Logo placeholder
    logo_tbl = Table(
        [
            [
                Paragraph(
                    "<b>TCA</b>",
                    ParagraphStyle(
                        "logo",
                        parent=styles["Normal"],
                        fontSize=24,
                        fontName="Helvetica-Bold",
                        textColor=colors.white,
                        alignment=1,
                    ),
                )
            ]
        ],
        colWidths=[2 * inch],
        rowHeights=[0.7 * inch],
    )
    logo_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), TCA_NAVY),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROUNDEDCORNERS", [6, 6, 6, 6]),
            ]
        )
    )
    logo_tbl.hAlign = "CENTER"
    story.append(logo_tbl)
    story.append(Spacer(1, 0.5 * inch))

    story.append(Paragraph("APEX Estimate Report", style_cover_title))
    story.append(HRFlowable(width="60%", thickness=2, color=TCA_NAVY, spaceAfter=16, hAlign="CENTER"))
    story.append(Paragraph(project.name, style_cover_sub))
    story.append(Spacer(1, 12))

    org_name = org.name if org else "General Contractor"
    cover_meta_lines = [
        f"<b>Prepared for:</b> {org_name}",
        f"<b>Project #:</b> {project.project_number}",
        f"<b>Location:</b> {project.location or '—'}",
        f"<b>Type:</b> {(project.project_type or '').title()}",
        f"<b>Date:</b> {date.today().strftime('%B %d, %Y')}",
        f"<b>Prepared by:</b> {current_user.full_name}",
    ]
    for line in cover_meta_lines:
        story.append(Paragraph(line, style_cover_meta))
        story.append(Spacer(1, 3))

    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════
    # PAGE 2: EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════
    story.append(Paragraph("Executive Summary", style_section))
    story.append(HRFlowable(width="100%", thickness=1, color=TCA_NAVY, spaceAfter=10))

    # Key metrics cards as a table
    line_items = estimate.line_items or []
    gap_count = gap_report.total_gaps if gap_report else 0
    confidence = None
    if gap_report and gap_report.overall_score is not None:
        confidence = f"{gap_report.overall_score:.0f}%"

    metrics = [
        ("Total Estimate", fmt_dollar(estimate.total_bid_amount)),
        ("Line Items", str(len(line_items))),
        ("Scope Confidence", confidence or "N/A"),
        ("Gaps Identified", str(gap_count)),
    ]
    metric_cells = []
    for label, value in metrics:
        cell = Paragraph(
            f'<font size="8" color="#64748B">{label}</font><br/><font size="16"><b>{value}</b></font>',
            ParagraphStyle("metric", parent=styles["Normal"], alignment=1, leading=20),
        )
        metric_cells.append(cell)

    card_w = usable_w / 4
    metric_tbl = Table([metric_cells], colWidths=[card_w] * 4, rowHeights=[0.8 * inch])
    metric_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GRAY),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.5, GRID_COLOR),
                ("LINEBEFORE", (1, 0), (1, 0), 0.5, GRID_COLOR),
                ("LINEBEFORE", (2, 0), (2, 0), 0.5, GRID_COLOR),
                ("LINEBEFORE", (3, 0), (3, 0), 0.5, GRID_COLOR),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(metric_tbl)
    story.append(Spacer(1, 16))

    # Executive summary narrative
    if estimate.executive_summary and estimate.executive_summary.strip():
        for para in estimate.executive_summary.split("\n\n"):
            if para.strip():
                story.append(Paragraph(para.strip(), style_body))
                story.append(Spacer(1, 4))
    else:
        story.append(Paragraph("No executive summary available for this estimate.", style_small))
    story.append(Spacer(1, 12))

    # Cost breakdown summary table
    story.append(Paragraph("Cost Summary", style_section))

    summary_rows = [
        [Paragraph("<b>Item</b>", styles["Normal"]), Paragraph("<b>Amount</b>", styles["Normal"])],
        ["Subtotal (Direct Cost)", fmt_dollar(estimate.total_direct_cost)],
        ["  Labor", fmt_dollar(estimate.total_labor_cost)],
        ["  Materials", fmt_dollar(estimate.total_material_cost)],
        ["  Subcontractor", fmt_dollar(estimate.total_subcontractor_cost)],
        [f"Overhead ({fmt_pct(estimate.overhead_pct)})", fmt_dollar(estimate.overhead_amount)],
        [f"Profit ({fmt_pct(estimate.profit_pct)})", fmt_dollar(estimate.profit_amount)],
        [f"Contingency ({fmt_pct(estimate.contingency_pct)})", fmt_dollar(estimate.contingency_amount)],
    ]
    if estimate.bid_bond_required:
        summary_rows.append(["Bond", "Included"])
    summary_rows.append(
        [
            Paragraph("<b>GRAND TOTAL BID</b>", styles["Normal"]),
            Paragraph(f"<b>{fmt_dollar(estimate.total_bid_amount)}</b>", styles["Normal"]),
        ]
    )

    sum_tbl = Table(summary_rows, colWidths=[4.0 * inch, 1.5 * inch])
    grand_row = len(summary_rows) - 1
    sum_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), TCA_NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, grand_row - 1), [colors.white, LIGHT_GRAY]),
                ("BACKGROUND", (0, grand_row), (-1, grand_row), TCA_NAVY_DARK),
                ("TEXTCOLOR", (0, grand_row), (-1, grand_row), colors.white),
                ("FONTNAME", (0, grand_row), (-1, grand_row), "Helvetica-Bold"),
                ("FONTSIZE", (0, grand_row), (-1, grand_row), 11),
                ("GRID", (0, 0), (-1, -1), 0.25, GRID_COLOR),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(sum_tbl)

    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════
    # PAGE 3+: LINE ITEMS TABLE
    # ══════════════════════════════════════════════════════════
    story.append(Paragraph("Estimate Line Items", style_section))
    story.append(HRFlowable(width="100%", thickness=1, color=TCA_NAVY, spaceAfter=10))

    # Group by division for subtotals
    divisions = {}
    for li in line_items:
        div = li.division_number or "00"
        if div not in divisions:
            divisions[div] = []
        divisions[div].append(li)

    col_widths = [0.5 * inch, 0.75 * inch, 2.3 * inch, 0.55 * inch, 0.45 * inch, 0.75 * inch, 0.75 * inch]
    header_row = [
        Paragraph("<b>Div</b>", styles["Normal"]),
        Paragraph("<b>CSI Code</b>", styles["Normal"]),
        Paragraph("<b>Description</b>", styles["Normal"]),
        Paragraph("<b>Qty</b>", styles["Normal"]),
        Paragraph("<b>Unit</b>", styles["Normal"]),
        Paragraph("<b>Unit Cost</b>", styles["Normal"]),
        Paragraph("<b>Total</b>", styles["Normal"]),
    ]

    tbl_data = [header_row]
    tbl_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), TCA_NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.25, GRID_COLOR),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    row_idx = 1
    grand_total = 0
    for div_num in sorted(divisions.keys()):
        items = divisions[div_num]
        div_total = 0
        for li in items:
            qty = li.quantity or 0
            unit_cost = li.unit_cost or 0
            total = li.total_cost or 0
            div_total += total
            tbl_data.append(
                [
                    li.division_number or "",
                    li.csi_code or "",
                    li.description or "",
                    f"{qty:,.1f}",
                    li.unit_of_measure or "",
                    fmt_dollar(unit_cost),
                    fmt_dollar(total),
                ]
            )
            row_idx += 1

        grand_total += div_total
        # Subtotal row
        tbl_data.append(
            [
                "",
                "",
                Paragraph(f"<b>Division {div_num} Subtotal</b>", styles["Normal"]),
                "",
                "",
                "",
                Paragraph(f"<b>{fmt_dollar(div_total)}</b>", styles["Normal"]),
            ]
        )
        tbl_style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), SUBTOTAL_BG))
        tbl_style_cmds.append(("FONTNAME", (0, row_idx), (-1, row_idx), "Helvetica-Bold"))
        row_idx += 1

    # Grand total row
    tbl_data.append(
        [
            "",
            "",
            Paragraph("<b>GRAND TOTAL (Direct)</b>", styles["Normal"]),
            "",
            "",
            "",
            Paragraph(f"<b>{fmt_dollar(grand_total)}</b>", styles["Normal"]),
        ]
    )
    tbl_style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), TCA_NAVY_DARK))
    tbl_style_cmds.append(("TEXTCOLOR", (0, row_idx), (-1, row_idx), colors.white))
    tbl_style_cmds.append(("FONTNAME", (0, row_idx), (-1, row_idx), "Helvetica-Bold"))

    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(tbl_style_cmds))
    story.append(tbl)
    story.append(Spacer(1, 16))

    # ── Exclusions & Assumptions ──
    if estimate.exclusions:
        story.append(Paragraph("Exclusions", style_section))
        for ex in estimate.exclusions:
            story.append(Paragraph(f"\u2022 {ex}", style_body))
        story.append(Spacer(1, 8))

    if estimate.assumptions:
        story.append(Paragraph("Assumptions", style_section))
        for asm in estimate.assumptions:
            story.append(Paragraph(f"\u2022 {asm}", style_body))
        story.append(Spacer(1, 8))

    # ══════════════════════════════════════════════════════════
    # GAP ANALYSIS SUMMARY
    # ══════════════════════════════════════════════════════════
    if gap_report and gap_report.items:
        story.append(PageBreak())
        story.append(Paragraph("Gap Analysis Summary", style_section))
        story.append(HRFlowable(width="100%", thickness=1, color=TCA_NAVY, spaceAfter=10))

        gap_summary_text = (
            f"Total gaps: {gap_report.total_gaps} &nbsp;|&nbsp; "
            f"Critical: {gap_report.critical_count} &nbsp;|&nbsp; "
            f"Moderate: {gap_report.moderate_count} &nbsp;|&nbsp; "
            f"Watch: {gap_report.watch_count}"
        )
        story.append(Paragraph(gap_summary_text, style_body))
        story.append(Spacer(1, 8))

        # Filter to critical/moderate gaps for the table
        significant_gaps = [g for g in gap_report.items if g.severity in ("critical", "moderate")]

        if significant_gaps:
            gap_col_widths = [0.7 * inch, 1.0 * inch, 2.0 * inch, 2.35 * inch]
            gap_header = [
                Paragraph("<b>CSI</b>", styles["Normal"]),
                Paragraph("<b>Severity</b>", styles["Normal"]),
                Paragraph("<b>Description</b>", styles["Normal"]),
                Paragraph("<b>Recommendation</b>", styles["Normal"]),
            ]
            gap_data = [gap_header]
            for g in significant_gaps:
                sev_color = "#DC2626" if g.severity == "critical" else "#D97706"
                gap_data.append(
                    [
                        g.section_number or g.division_number or "",
                        Paragraph(
                            f'<font color="{sev_color}"><b>{(g.severity or "").title()}</b></font>', styles["Normal"]
                        ),
                        Paragraph(g.description or g.title or "", style_small),
                        Paragraph(g.recommendation or "—", style_small),
                    ]
                )

            gap_tbl = Table(gap_data, colWidths=gap_col_widths, repeatRows=1)
            gap_tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), TCA_NAVY),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
                        ("GRID", (0, 0), (-1, -1), 0.25, GRID_COLOR),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            story.append(gap_tbl)
        story.append(Spacer(1, 12))

    # ══════════════════════════════════════════════════════════
    # METHODOLOGY NOTE
    # ══════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("Methodology", style_section))
    story.append(HRFlowable(width="100%", thickness=1, color=TCA_NAVY, spaceAfter=10))

    methodology_text = (
        "This estimate was generated by the APEX Estimating Platform, a 7-agent AI pipeline. "
        "Documents are ingested and parsed (Agent 1), scope sections are extracted with quantities "
        "(Agent 2), gaps are identified against a master CSI checklist (Agent 3), quantities are "
        "verified via takeoff analysis (Agent 4), labor and productivity rates are matched from "
        "historical data (Agent 5), and the final estimate is assembled with deterministic Python "
        "math (Agent 6). Agent 7 provides variance analysis against project actuals when available."
    )
    story.append(Paragraph(methodology_text, style_body))
    story.append(Spacer(1, 8))

    story.append(
        Paragraph(
            "<b>Important:</b> All dollar amounts in this report are computed using deterministic "
            "Python arithmetic. No LLM generates or modifies financial figures. AI is used for "
            "document parsing, scope analysis, and rate matching only.",
            style_body,
        )
    )
    story.append(Spacer(1, 8))

    story.append(
        Paragraph(
            "Data sources include uploaded project specifications, the APEX productivity library "
            "(historical labor rates), material price databases, and benchmark data from completed "
            "projects within the organization.",
            style_body,
        )
    )
    story.append(Spacer(1, 16))

    # ══════════════════════════════════════════════════════════
    # APPENDIX: AGENT RUN LOG
    # ══════════════════════════════════════════════════════════
    if agent_logs:
        story.append(Paragraph("Appendix: Agent Pipeline Run Log", style_section))
        story.append(HRFlowable(width="100%", thickness=1, color=TCA_NAVY, spaceAfter=10))

        log_col_widths = [0.4 * inch, 1.6 * inch, 0.7 * inch, 0.8 * inch, 0.8 * inch, 1.75 * inch]
        log_header = [
            Paragraph("<b>#</b>", styles["Normal"]),
            Paragraph("<b>Agent</b>", styles["Normal"]),
            Paragraph("<b>Status</b>", styles["Normal"]),
            Paragraph("<b>Duration</b>", styles["Normal"]),
            Paragraph("<b>Tokens</b>", styles["Normal"]),
            Paragraph("<b>Summary</b>", styles["Normal"]),
        ]
        log_data = [log_header]
        for log in agent_logs:
            status_color = "#16A34A" if log.status == "completed" else "#DC2626" if log.status == "error" else "#64748B"
            duration = f"{log.duration_seconds:.1f}s" if log.duration_seconds else "—"
            tokens = f"{log.tokens_used:,}" if log.tokens_used else "—"
            summary = (log.output_summary or "")[:80]
            if len(log.output_summary or "") > 80:
                summary += "…"

            log_data.append(
                [
                    str(log.agent_number),
                    log.agent_name or "",
                    Paragraph(
                        f'<font color="{status_color}"><b>{(log.status or "").title()}</b></font>', styles["Normal"]
                    ),
                    duration,
                    tokens,
                    Paragraph(summary or "—", style_small),
                ]
            )

        log_tbl = Table(log_data, colWidths=log_col_widths, repeatRows=1)
        log_tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), TCA_NAVY),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
                    ("GRID", (0, 0), (-1, -1), 0.25, GRID_COLOR),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(log_tbl)
        story.append(Spacer(1, 12))

    # ── Final footer ──
    story.append(HRFlowable(width="100%", thickness=1, color=MID_GRAY, spaceBefore=16, spaceAfter=6))
    story.append(
        Paragraph(
            f"Prepared by {current_user.full_name} &nbsp;|&nbsp; Generated by APEX Estimating Platform &nbsp;|&nbsp; {date.today().strftime('%B %d, %Y')}",
            style_footer_text,
        )
    )

    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    buf.seek(0)

    filename = f"{project.project_number}_estimate.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────── XLSX ───────────────────────────


@router.get("/projects/{project_id}/estimate/xlsx")
def export_estimate_xlsx(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    project, estimate, org = _get_estimate_or_404(project_id, db, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    APEX_BLUE = "1E40AF"
    LIGHT_BLUE = "DBEAFE"
    LIGHT_GRAY = "F1F5F9"
    DARK_BLUE = "1E3A8A"
    WHITE = "FFFFFF"

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(color="FFFFFF", size=10, bold=True):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def cell_font(bold=False, size=9):
        return Font(name="Calibri", bold=bold, size=size)

    def fill(hex_color):
        return PatternFill(fill_type="solid", fgColor=hex_color)

    def currency_fmt(ws, row, col):
        ws.cell(row=row, column=col).number_format = '"$"#,##0.00'

    # ── Rows 1-3: Header info ──
    org_name = org.name if org else "General Contractor"
    ws["A1"] = org_name
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=APEX_BLUE)

    ws["A2"] = project.name
    ws["A2"].font = Font(name="Calibri", bold=True, size=12)

    ws["A3"] = (
        f"Project #: {project.project_number}  |  "
        f"Location: {project.location or '—'}  |  "
        f"Date: {date.today().strftime('%B %d, %Y')}"
    )
    ws["A3"].font = Font(name="Calibri", size=9, color="64748B")

    # Merge header cells across table width (8 cols)
    for row in [1, 2, 3]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    ws.row_dimensions[4].height = 6  # blank spacer

    # ── Row 5: Column headers ──
    headers = ["Item #", "CSI Code", "Description", "Quantity", "Unit", "Unit Cost", "Total", "Notes"]
    for col, hdr in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=hdr)
        c.font = hdr_font()
        c.fill = fill(APEX_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    ws.row_dimensions[5].height = 18

    # ── Data rows (starting at row 6) ──
    line_items = estimate.line_items or []

    # Group by division
    divisions = {}
    for li in line_items:
        div = li.division_number or "00"
        if div not in divisions:
            divisions[div] = []
        divisions[div].append(li)

    current_row = 6
    item_num = 1
    data_start_rows = {}  # div -> (first_data_row, last_data_row)

    row_bg = [LIGHT_GRAY, WHITE]

    for div_num in sorted(divisions.keys()):
        items = divisions[div_num]
        div_first = current_row

        for li in items:
            bg = row_bg[(current_row - 6) % 2]
            ws.cell(row=current_row, column=1, value=item_num).font = cell_font()
            ws.cell(row=current_row, column=2, value=li.csi_code or "").font = cell_font()
            ws.cell(row=current_row, column=3, value=li.description or "").font = cell_font()
            ws.cell(row=current_row, column=4, value=li.quantity or 0).font = cell_font()
            ws.cell(row=current_row, column=5, value=li.unit_of_measure or "").font = cell_font()
            ws.cell(row=current_row, column=6, value=li.unit_cost or 0).font = cell_font()
            ws.cell(row=current_row, column=7, value=li.total_cost or 0).font = cell_font()
            ws.cell(row=current_row, column=8, value=li.notes or "").font = cell_font()

            ws.cell(row=current_row, column=6).number_format = '"$"#,##0.00'
            ws.cell(row=current_row, column=7).number_format = '"$"#,##0.00'

            for col in range(1, 9):
                c = ws.cell(row=current_row, column=col)
                c.fill = fill(bg)
                c.border = border
                c.alignment = Alignment(vertical="center")

            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal="right", vertical="center")

            item_num += 1
            current_row += 1

        data_start_rows[div_num] = (div_first, current_row - 1)

        # Division subtotal row with SUM formula
        div_last = current_row - 1
        subtotal_col_g = f"G{div_first}:G{div_last}"
        ws.cell(row=current_row, column=3, value=f"Division {div_num} Subtotal").font = hdr_font(
            color=APEX_BLUE, size=9
        )
        ws.cell(row=current_row, column=7, value=f"=SUM({subtotal_col_g})").font = hdr_font(color=APEX_BLUE, size=9)
        ws.cell(row=current_row, column=7).number_format = '"$"#,##0.00'

        for col in range(1, 9):
            c = ws.cell(row=current_row, column=col)
            c.fill = fill(LIGHT_BLUE)
            c.border = border
            c.alignment = Alignment(vertical="center")
        ws.cell(row=current_row, column=7).alignment = Alignment(horizontal="right", vertical="center")

        current_row += 1

    # ── Summary section ──
    current_row += 1  # blank row
    summary_start = current_row

    def sum_row(label, formula_or_val, is_pct=False, bold=False, bg=WHITE, text_color="000000"):
        ws.cell(row=current_row, column=6, value=label).font = Font(name="Calibri", bold=bold, size=9, color=text_color)
        c = ws.cell(row=current_row, column=7, value=formula_or_val)
        c.font = Font(name="Calibri", bold=bold, size=9, color=text_color)
        if not is_pct:
            c.number_format = '"$"#,##0.00'
        c.alignment = Alignment(horizontal="right", vertical="center")
        for col in [6, 7]:
            ws.cell(row=current_row, column=col).fill = fill(bg)
            ws.cell(row=current_row, column=col).border = border

    # Collect all division subtotal rows to build the overall subtotal formula
    div_subtotal_row_refs = []
    row_scan = 6
    while row_scan < summary_start:
        val = ws.cell(row=row_scan, column=3).value
        if val and isinstance(val, str) and "Subtotal" in val:
            div_subtotal_row_refs.append(f"G{row_scan}")
        row_scan += 1

    subtotal_formula = "=" + "+".join(div_subtotal_row_refs) if div_subtotal_row_refs else "=0"

    overhead_pct = (estimate.overhead_pct or 0) / 100
    profit_pct = (estimate.profit_pct or 0) / 100
    contingency_pct = (estimate.contingency_pct or 0) / 100

    subtotal_ref = f"G{current_row}"
    sum_row("Subtotal (Direct Cost)", subtotal_formula, bg=LIGHT_GRAY, bold=True)
    current_row += 1

    overhead_ref = f"G{current_row}"
    sum_row(f"Overhead ({estimate.overhead_pct or 0:.1f}%)", f"={subtotal_ref}*{overhead_pct}", bg=LIGHT_GRAY)
    current_row += 1

    profit_ref = f"G{current_row}"
    sum_row(
        f"Profit ({estimate.profit_pct or 0:.1f}%)", f"=({subtotal_ref}+{overhead_ref})*{profit_pct}", bg=LIGHT_GRAY
    )
    current_row += 1

    contingency_ref = f"G{current_row}"
    sum_row(
        f"Contingency ({estimate.contingency_pct or 0:.1f}%)",
        f"=({subtotal_ref}+{overhead_ref}+{profit_ref})*{contingency_pct}",
        bg=LIGHT_GRAY,
    )
    current_row += 1

    sum_row(
        "GRAND TOTAL BID",
        f"={subtotal_ref}+{overhead_ref}+{profit_ref}+{contingency_ref}",
        bold=True,
        bg=DARK_BLUE,
        text_color=WHITE,
    )
    for col in [6, 7]:
        ws.cell(row=current_row, column=col).font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    current_row += 1

    # ── Column widths ──
    col_widths = [7, 12, 45, 10, 8, 14, 14, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ──
    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{project.project_number}_estimate.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────── SUBCONTRACTOR PACKAGES ────────────────────────

# Map CSI division prefix → trade name
_TRADE_MAP = {
    "01": "General Conditions",
    "02": "Existing Conditions",
    "03": "Concrete",
    "04": "Masonry",
    "05": "Metals",
    "06": "Wood & Plastics",
    "07": "Thermal & Moisture",
    "08": "Openings",
    "09": "Finishes",
    "10": "Specialties",
    "11": "Equipment",
    "12": "Furnishings",
    "13": "Special Construction",
    "14": "Conveying",
    "21": "Fire Suppression",
    "22": "Plumbing",
    "23": "HVAC",
    "25": "Integrated Automation",
    "26": "Electrical",
    "27": "Communications",
    "28": "Electronic Safety",
    "31": "Earthwork",
    "32": "Exterior Improvements",
    "33": "Utilities",
}


@router.get("/projects/{project_id}/subcontractor-packages/list")
def list_subcontractor_packages(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Return the trade breakdown without generating PDFs — for the UI table."""
    project, estimate, org = _get_estimate_or_404(project_id, db)

    trades: dict[str, dict] = {}
    for li in estimate.line_items or []:
        div = (li.division_number or "00").strip()
        trade = _TRADE_MAP.get(div[:2] if len(div) >= 2 else div, f"Division {div}")
        if trade not in trades:
            trades[trade] = {"division": div[:2] if len(div) >= 2 else div, "items": 0, "total": 0.0}
        trades[trade]["items"] += 1
        trades[trade]["total"] += li.total_cost or 0.0

    result = [
        {"trade": t, "division": v["division"], "items": v["items"], "total": v["total"]}
        for t, v in sorted(trades.items(), key=lambda x: x[1]["division"])
    ]
    return {"success": True, "data": result}


@router.get("/projects/{project_id}/subcontractor-packages/{trade}")
def export_subcontractor_package_pdf(
    project_id: int,
    trade: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Generate a PDF bid package for a single trade / CSI division."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    project, estimate, org = _get_estimate_or_404(project_id, db)

    trade_decoded = trade.replace("-", " ").title()

    # Collect matching line items (match by trade name or division prefix)
    div_prefix = next(
        (k for k, v in _TRADE_MAP.items() if v.lower() == trade_decoded.lower()),
        None,
    )
    line_items = [
        li
        for li in (estimate.line_items or [])
        if (li.division_number or "").startswith(div_prefix or "__NOMATCH__")
        or (li.division_number or "") == trade_decoded
    ]

    if not line_items:
        raise HTTPException(
            status_code=404,
            detail=f"No line items found for trade '{trade_decoded}'",
        )

    buf = io.BytesIO()
    styles = getSampleStyleSheet()
    APEX_BLUE = colors.HexColor("#1e40af")
    LIGHT_GRAY = colors.HexColor("#f1f5f9")
    MID_GRAY = colors.HexColor("#94a3b8")

    style_title = ParagraphStyle(
        "title", parent=styles["Normal"], fontSize=18, fontName="Helvetica-Bold", textColor=APEX_BLUE, spaceAfter=4
    )
    style_sub = ParagraphStyle(
        "sub",
        parent=styles["Normal"],
        fontSize=11,
        fontName="Helvetica-Bold",
        textColor=APEX_BLUE,
        spaceBefore=10,
        spaceAfter=4,
    )
    style_meta = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#475569"))
    style_footer = ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=MID_GRAY, alignment=1)

    def fmt_dollar(val):
        return f"${val:,.0f}" if val is not None else "$0"

    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    story = []

    org_name = org.name if org else "General Contractor"
    story.append(
        Paragraph(org_name.upper(), ParagraphStyle("org", parent=styles["Normal"], fontSize=9, textColor=MID_GRAY))
    )
    story.append(Spacer(1, 4))
    story.append(Paragraph("SUBCONTRACTOR BID PACKAGE", style_title))
    story.append(Paragraph(f"{project.name} — {trade_decoded}", style_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=APEX_BLUE, spaceAfter=6))

    meta = (
        f"<b>Project #:</b> {project.project_number} &nbsp;&nbsp; "
        f"<b>Location:</b> {project.location or '—'} &nbsp;&nbsp; "
        f"<b>Bid Date:</b> {project.bid_date or '—'} &nbsp;&nbsp; "
        f"<b>Date Issued:</b> {date.today().strftime('%B %d, %Y')}"
    )
    story.append(Paragraph(meta, style_meta))
    story.append(Spacer(1, 16))

    story.append(Paragraph("Scope of Work", style_sub))
    story.append(
        Paragraph(
            f"The following scope of work is to be performed by the subcontractor for "
            f"<b>{trade_decoded}</b> work. All work shall conform to project specifications, "
            f"applicable codes, and general contractor requirements.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 12))

    story.append(Paragraph("Bid Items", style_sub))
    col_widths = [0.7 * inch, 2.8 * inch, 0.6 * inch, 0.5 * inch, 0.9 * inch, 0.9 * inch]
    header_row = [
        Paragraph("<b>CSI Code</b>", styles["Normal"]),
        Paragraph("<b>Description</b>", styles["Normal"]),
        Paragraph("<b>Qty</b>", styles["Normal"]),
        Paragraph("<b>Unit</b>", styles["Normal"]),
        Paragraph("<b>Unit Cost</b>", styles["Normal"]),
        Paragraph("<b>Total</b>", styles["Normal"]),
    ]
    tbl_data = [header_row]
    total_cost = 0.0

    for li in line_items:
        tbl_data.append(
            [
                li.csi_code or "",
                li.description or "",
                f"{li.quantity or 0:,.1f}",
                li.unit_of_measure or "",
                fmt_dollar(li.unit_cost),
                fmt_dollar(li.total_cost),
            ]
        )
        total_cost += li.total_cost or 0.0

    tbl_data.append(
        [
            "",
            Paragraph("<b>SUBTOTAL</b>", styles["Normal"]),
            "",
            "",
            "",
            Paragraph(f"<b>{fmt_dollar(total_cost)}</b>", styles["Normal"]),
        ]
    )

    tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    last = len(tbl_data) - 1
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), APEX_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, last - 1), [colors.white, LIGHT_GRAY]),
                ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#dbeafe")),
                ("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 20))

    story.append(Paragraph("Terms & Conditions", style_sub))
    for term in [
        "Subcontractor shall provide all labor, material, equipment, and supervision.",
        "All work subject to general contractor review and approval.",
        "Subcontractor shall carry liability insurance per project requirements.",
        "Submission of bid constitutes acceptance of project schedule and milestones.",
        "Unit prices shall remain firm for 30 days from bid date.",
    ]:
        story.append(Paragraph(f"• {term}", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(HRFlowable(width="100%", thickness=1, color=MID_GRAY, spaceBefore=12, spaceAfter=6))
    story.append(
        Paragraph(
            f"Package prepared by {current_user.full_name} | APEX Estimating Platform",
            style_footer,
        )
    )

    doc.build(story)
    buf.seek(0)

    safe_trade = trade_decoded.replace(" ", "-").replace("/", "-").lower()
    filename = f"{project.project_number}_{safe_trade}_bid_package.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────── CSV ───────────────────────────


@router.get("/projects/{project_id}/estimate/csv")
def export_estimate_csv(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Export estimate as CSV for accounting import."""
    import csv

    project, estimate, org = _get_estimate_or_404(project_id, db, current_user)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Item #",
            "CSI Code",
            "Division",
            "Description",
            "Quantity",
            "UOM",
            "Labor Cost",
            "Material Cost",
            "Equipment Cost",
            "Subcontractor Cost",
            "Total Cost",
            "Unit Cost",
        ]
    )
    for idx, li in enumerate(estimate.line_items, 1):
        writer.writerow(
            [
                idx,
                li.csi_code,
                li.division_number,
                li.description,
                li.quantity,
                li.unit_of_measure,
                f"{li.labor_cost:.2f}",
                f"{li.material_cost:.2f}",
                f"{li.equipment_cost:.2f}",
                f"{li.subcontractor_cost:.2f}",
                f"{li.total_cost:.2f}",
                f"{li.unit_cost:.2f}",
            ]
        )
    # Summary rows
    writer.writerow([])
    writer.writerow(
        ["", "", "", "Subtotal (Direct Cost)", "", "", "", "", "", "", f"{estimate.total_direct_cost:.2f}", ""]
    )
    writer.writerow(
        [
            "",
            "",
            "",
            f"Overhead ({estimate.overhead_pct}%)",
            "",
            "",
            "",
            "",
            "",
            "",
            f"{estimate.overhead_amount:.2f}",
            "",
        ]
    )
    writer.writerow(
        ["", "", "", f"Profit ({estimate.profit_pct}%)", "", "", "", "", "", "", f"{estimate.profit_amount:.2f}", ""]
    )
    writer.writerow(
        [
            "",
            "",
            "",
            f"Contingency ({estimate.contingency_pct}%)",
            "",
            "",
            "",
            "",
            "",
            "",
            f"{estimate.contingency_amount:.2f}",
            "",
        ]
    )
    writer.writerow(["", "", "", "GRAND TOTAL", "", "", "", "", "", "", f"{estimate.total_bid_amount:.2f}", ""])

    filename = f"{project.project_number}_estimate.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────── QuickBooks IIF ───────────────────────────


@router.get("/projects/{project_id}/estimate/qb")
def export_estimate_quickbooks(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Export estimate in QuickBooks IIF format."""
    project, estimate, org = _get_estimate_or_404(project_id, db, current_user)

    lines = []
    lines.append("!TRNS\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO")
    lines.append("!SPL\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO")
    lines.append("!ENDTRNS")

    today = date.today().strftime("%m/%d/%Y")

    # Main transaction
    lines.append(
        f"TRNS\tESTIMATE\t{today}\tAccounts Receivable\t{project.name}\t{estimate.total_bid_amount:.2f}\tProject Estimate {project.project_number}"
    )

    for li in estimate.line_items:
        lines.append(
            f"SPL\tESTIMATE\t{today}\tConstruction Income:{li.division_number}\t{project.name}\t-{li.total_cost:.2f}\t{li.csi_code} {li.description[:50]}"
        )

    # Markup splits
    if estimate.overhead_amount > 0:
        lines.append(
            f"SPL\tESTIMATE\t{today}\tOverhead\t{project.name}\t-{estimate.overhead_amount:.2f}\tOverhead {estimate.overhead_pct}%"
        )
    if estimate.profit_amount > 0:
        lines.append(
            f"SPL\tESTIMATE\t{today}\tProfit\t{project.name}\t-{estimate.profit_amount:.2f}\tProfit {estimate.profit_pct}%"
        )
    if estimate.contingency_amount > 0:
        lines.append(
            f"SPL\tESTIMATE\t{today}\tContingency\t{project.name}\t-{estimate.contingency_amount:.2f}\tContingency {estimate.contingency_pct}%"
        )

    lines.append("ENDTRNS")

    content = "\n".join(lines)
    filename = f"{project.project_number}_estimate.iif"

    return StreamingResponse(
        iter([content]),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
